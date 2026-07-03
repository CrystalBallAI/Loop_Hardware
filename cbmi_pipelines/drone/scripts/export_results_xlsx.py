#!/usr/bin/env python3
"""Export Drone PPK pipeline results to a single Excel workbook.

SINGLE-RECORD layout (one survey = one drone flight). Eight sheets total:
  Summary | 1_Source_Files | 2_Source_Fields | 3_Derived_Fields | 4_Indicators
  5_Building_Blocks (drone_score blocks only) | 5b_CAL_CONF (parallel, non-contributing)
  6_Drone_Score

VALUES ONLY — no Excel formulas. The numbers are the authoritative pipeline
results and the workbook must render correctly with no recalc engine.

Run: python3 scripts/export_results_xlsx.py paths.json
"""
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F3864")          # dark navy
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
SUBHEAD_FILL = PatternFill("solid", fgColor="305496")         # mid navy
SUBHEAD_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name=FONT_NAME, size=10)
BODY_BOLD = Font(name=FONT_NAME, bold=True, size=10)

NA_FILL = PatternFill("solid", fgColor="F2F2F2")              # light grey for null rows
FLAG_FILL = PatternFill("solid", fgColor="FCE4D6")            # peach for flagged rows
GATE_FILL = PatternFill("solid", fgColor="F8CBAD")            # darker peach for gate cells
CROSSDOC_FILL = PatternFill("solid", fgColor="FFF2CC")        # pale yellow for handoff
APEX_FILL = PatternFill("solid", fgColor="C6E0B4")            # pale green for apex / sum rows
CAL_CONF_FILL = PatternFill("solid", fgColor="E4D7EC")        # lavender for the 5b sheet body
CAL_CONF_BANNER_FILL = PatternFill("solid", fgColor="B59ED1") # deeper lavender for banner

THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

# Recommendations color palette (per RECOMMENDATIONS_BUILD_PROMPT STEP D)
LEVEL_FILLS = {
    "good":     PatternFill("solid", fgColor="E2EFDA"),  # light green
    "minor":    PatternFill("solid", fgColor="DDEBF7"),  # light blue
    "review":   PatternFill("solid", fgColor="FFF2CC"),  # light yellow
    "resurvey": PatternFill("solid", fgColor="FCE4D6"),  # peach
    "unknown":  PatternFill("solid", fgColor="E7E6E6"),  # light grey
}
DECISION_FILLS = {
    "good_to_go":             PatternFill("solid", fgColor="70AD47"),  # green
    "review_recommended":     PatternFill("solid", fgColor="FFC000"),  # amber
    "resurvey_recommended":   PatternFill("solid", fgColor="C00000"),  # red
    "unable_to_assess":       PatternFill("solid", fgColor="808080"),  # grey
}
DECISION_FG = {
    "good_to_go":             "FFFFFF",
    "review_recommended":     "000000",
    "resurvey_recommended":   "FFFFFF",
    "unable_to_assess":       "FFFFFF",
}


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def stringify(v, limit: int = 200) -> str:
    """Compress dict/list to JSON, truncate."""
    if v is None:
        return ""
    if isinstance(v, (dict, list, tuple)):
        s = json.dumps(v, sort_keys=True, default=str)
    elif isinstance(v, bool):
        s = "true" if v else "false"
    else:
        s = str(v)
    if len(s) > limit:
        s = s[: limit - 3] + "..."
    return s


def style_header_row(ws, row_idx: int):
    for cell in ws[row_idx]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        cell.border = THIN_BORDER


def style_body_cell(cell, fill=None, bold=False):
    cell.font = BODY_BOLD if bold else BODY_FONT
    if fill is not None:
        cell.fill = fill
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)


def set_column_widths(ws, widths: list):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def freeze_header_and_id_col(ws, header_row: int = 1, id_col: int = 2):
    """Freeze the header row + the first id column so values scroll under fixed labels."""
    ws.freeze_panes = ws.cell(row=header_row + 1, column=id_col + 1)


# ---------------------------------------------------------------------------
# sheet writers
# ---------------------------------------------------------------------------

def write_summary(ws, *, apex: dict, spec: dict, blocks: dict, cal_conf: dict, sheet_legend: list, rec=None):
    apex_score = apex.get("drone_score")
    gate_triggered = apex.get("global_gate_triggered")
    img_cap_score = apex["block_contributions"]["BB_IMG_CAPTURE"]["block_score"]
    flags_severity = apex.get("all_flags_by_severity", {})
    flags = apex.get("all_flags_aggregated", [])
    crossdoc = apex.get("_handoff_crossdoc_candidates", [])
    spec_version = spec["_meta"]["version"]
    cal_conf_score = cal_conf["score"] if cal_conf else None

    # ---- headline ----
    ws.append(["Drone PPK Pipeline — Results Summary"])
    ws["A1"].font = Font(name=FONT_NAME, bold=True, size=16, color="1F3864")
    ws.merge_cells("A1:E1")
    ws.append([])

    ws.append(["Apex score",        "drone_score",          apex_score, "out of 100"])
    ws.append(["Apex formula",      apex.get("formula_expression")])
    ws.append(["Spec version",      spec_version])
    ws.append(["Workflow",          apex.get("workflow")])
    ws.append(["Raw weighted sum",  apex.get("raw_weighted_sum")])
    ws.append([])

    # ---- Recommendation callout (only if 07_recommendations.json was loaded) ----
    if rec is not None:
        ws.append(["Recommendation"])
        style_header_row(ws, ws.max_row)
        decision = rec.get("decision", "unknown")
        decision_display = decision.upper().replace("_", " ")
        dec_row_idx = ws.max_row + 1
        ws.append(["Decision", decision_display])
        dec_cell = ws.cell(row=dec_row_idx, column=2)
        dec_cell.fill = DECISION_FILLS.get(decision, DECISION_FILLS["unable_to_assess"])
        dec_cell.font = Font(name=FONT_NAME, bold=True, size=13,
                             color=DECISION_FG.get(decision, "FFFFFF"))
        dec_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[dec_row_idx].height = 24
        ws.append(["Tier",                rec.get("tier_interpretation")])
        ws.append(["Rationale",           rec.get("decision_rationale")])
        s = rec.get("summary", {})
        ws.append(["Indicators by level",
                   f"good={s.get('good_count', 0)}  minor={s.get('minor_count', 0)}  "
                   f"review={s.get('review_count', 0)}  resurvey={s.get('resurvey_count', 0)}"])
        hg = s.get("hard_gates_fired", [])
        hg_text = ", ".join(hg) if hg else "(none)"
        hg_row_idx = ws.max_row + 1
        ws.append(["Hard gates fired", hg_text])
        if hg:
            ws.cell(row=hg_row_idx, column=2).fill = LEVEL_FILLS["resurvey"]
        ws.append(["Library version", rec.get("library_version")])
        if rec.get("_spec_version_note"):
            ws.append(["Spec/library note", rec["_spec_version_note"]])
        ws.append([])

    # ---- global gate ----
    ws.append(["Global gate"])
    style_header_row(ws, ws.max_row)
    ws.append(["triggered",            str(gate_triggered)])
    ws.append(["condition",            apex.get("global_gate_condition")])
    ws.append(["action",               apex.get("global_gate_action")])
    ws.append(["block_score_observed (BB_IMG_CAPTURE)", img_cap_score])
    ws.append([])

    # ---- flag summary by severity ----
    ws.append(["Flags by severity"])
    style_header_row(ws, ws.max_row)
    for sev in ("CRITICAL", "HIGH", "MEDIUM", "LOW"):
        names = flags_severity.get(sev, [])
        ws.append([sev, len(names), ", ".join(names) if names else "(none)"])
    ws.append([])

    # ---- flag summary by origin stage ----
    ws.append(["Flags by origin stage"])
    style_header_row(ws, ws.max_row)
    by_stage = {}
    for f in flags:
        by_stage.setdefault(f.get("_origin_stage", "(unknown)"), []).append(f.get("flag_name"))
    if not by_stage:
        ws.append(["(no flags raised)"])
    else:
        for stg, names in by_stage.items():
            ws.append([stg, len(names), ", ".join(names)])
    ws.append([])

    # ---- cross-doc handoff candidates ----
    ws.append(["Cross-document handoff candidates"])
    style_header_row(ws, ws.max_row)
    if not crossdoc:
        ws.append(["(none — apex envelope does not list any _handoff_crossdoc_candidates this run)"])
    else:
        ws.append(["flag_id", "flag_name", "target_subsystem", "context"])
        for c in crossdoc:
            ws.append([c.get("flag_id"), c.get("flag_name"), c.get("target_subsystem"),
                       stringify(c.get("context"))])
    ws.append([])

    # ---- CAL_CONF parallel deliverable callout ----
    ws.append(["CAL_CONF (parallel deliverable — does NOT contribute to drone_score)"])
    style_header_row(ws, ws.max_row)
    ws.append(["score", cal_conf_score, f"weight in drone_score = 0.00 (sheet 5b carries the full breakdown)"])
    ws.append([])

    # ---- sheet legend ----
    ws.append(["Sheets in this workbook"])
    style_header_row(ws, ws.max_row)
    for s in sheet_legend:
        ws.append(s)
    style_header_row(ws, ws.max_row - len(sheet_legend))

    # style body
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        for cell in row:
            if cell.font.bold and cell.fill.fgColor.rgb in (HEADER_FILL.fgColor.rgb,):
                continue
            cell.font = BODY_FONT if not cell.font.bold else BODY_BOLD
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)

    set_column_widths(ws, [44, 22, 60, 40, 30])


def write_source_files(ws, *, inventory: dict, spec: dict):
    src_files_spec = spec["source_files"]
    inv = inventory.get("data", inventory)

    # header
    ws.append(["file_id", "file_name", "subsystem", "category", "found", "physical_count", "notes"])
    style_header_row(ws, 1)

    # Build a probe over the spec's source files; cross-reference inventory's section blobs
    inv_map = {
        "SRC_IMG_01": inv.get("images", {}),
        "SRC_IMG_02": inv.get("calibration_library", {}),
        "SRC_GNSS_01": inv.get("rinex", {}),
        "SRC_FC_BIN": inv.get("bin", {}),
        "SRC_UI_01": inv.get("user_input", {}),
        "SRC_API_01": {"exists": True, "_meta": "Open-Meteo HTTPS"},
    }
    for s in src_files_spec:
        fid = s["file_id"]
        sec = inv_map.get(fid, {})
        if fid == "SRC_IMG_01":
            count, found = sec.get("file_count"), bool(sec.get("file_count"))
        elif fid == "SRC_GNSS_01":
            count = len(sec.get("observation_files", []) or []) + len(sec.get("navigation_files", []) or [])
            found = count > 0
        elif fid == "SRC_FC_BIN":
            count, found = len(sec.get("bin_files", []) or []), bool(sec.get("bin_files"))
        elif fid == "SRC_IMG_02":
            count, found = sec.get("entry_count"), bool(sec.get("exists"))
        elif fid == "SRC_UI_01":
            count = 1 if sec.get("exists") else 0
            found = bool(sec.get("exists"))
        elif fid == "SRC_API_01":
            count, found = 1, True
        else:
            count, found = None, None
        notes = (sec.get("hard_failure") or sec.get("warning") or "") or ""
        ws.append([fid, s["file_name"], s["subsystem"], s.get("file_extensions"),
                   "yes" if found else "no",
                   count if count is not None else "",
                   notes])

    # body styling
    for r in range(2, ws.max_row + 1):
        for c in ws[r]:
            style_body_cell(c)
        # paint not-found rows with a soft grey
        if ws.cell(row=r, column=5).value == "no":
            for c in ws[r]:
                c.fill = NA_FILL

    set_column_widths(ws, [14, 26, 26, 28, 8, 14, 50])
    freeze_header_and_id_col(ws, header_row=1, id_col=2)


def write_source_fields(ws, *, source_envelope: dict, spec: dict):
    data = source_envelope["data"]
    rows = []
    for f in spec["source_fields"]:
        fid = f["field_id"]
        value = data.get(fid)
        rows.append({
            "field_id": fid,
            "field_name": f["field_name"],
            "file_id": f["file_id"],
            "category": f.get("category"),
            "data_type": f.get("data_type"),
            "units": f.get("units") or "",
            "value": value,
        })

    ws.append(["field_id", "field_name", "file_id", "category", "data_type", "units", "value"])
    style_header_row(ws, 1)
    for r in rows:
        ws.append([
            r["field_id"], r["field_name"], r["file_id"],
            r["category"], r["data_type"], r["units"],
            stringify(r["value"]),
        ])

    # styling: grey row for null values
    for r in range(2, ws.max_row + 1):
        val_cell = ws.cell(row=r, column=7)
        is_empty = val_cell.value in ("", None, "null")
        for c in ws[r]:
            style_body_cell(c, fill=(NA_FILL if is_empty else None))

    set_column_widths(ws, [16, 28, 14, 24, 12, 16, 60])
    freeze_header_and_id_col(ws, header_row=1, id_col=2)


def write_derived_fields(ws, *, derived_envelope: dict, spec: dict):
    data = derived_envelope["data"]
    notes_blob = data.get("_notes", {}) or {}

    ws.append(["derived_id", "derived_name", "subsystem", "units", "value", "is_null", "note"])
    style_header_row(ws, 1)
    for d in spec["derived_fields"]:
        did = d["derived_id"]
        v = data.get(did)
        ws.append([
            did, d["derived_name"], d.get("subsystem"),
            d.get("units") or "",
            stringify(v),
            "yes" if v is None else "no",
            notes_blob.get(did, ""),
        ])

    for r in range(2, ws.max_row + 1):
        is_null = ws.cell(row=r, column=6).value == "yes"
        for c in ws[r]:
            style_body_cell(c, fill=(NA_FILL if is_null else None))

    set_column_widths(ws, [16, 32, 22, 16, 26, 8, 80])
    freeze_header_and_id_col(ws, header_row=1, id_col=2)


def write_indicators(ws, *, indicators_envelope: dict, spec: dict):
    traces = indicators_envelope["data"]["indicators"]
    by_id = {t["indicator_id"]: t for t in traces}

    ws.append(["indicator_id", "indicator_name", "block_id", "weight_in_block",
               "score", "band_matched", "condition", "input_value", "flags_raised"])
    style_header_row(ws, 1)

    for ind in spec["indicators"]:
        iid = ind["indicator_id"]
        t = by_id.get(iid, {})
        fls = t.get("flags_raised") or []
        flag_names = ", ".join(f.get("flag_name", "") for f in fls)
        ws.append([
            iid, t.get("indicator_name") or ind.get("indicator_name"),
            ind.get("building_block_id"), ind.get("weight_in_block"),
            t.get("score"),
            t.get("band_matched"),
            stringify(t.get("condition")),
            stringify(t.get("input_value")),
            flag_names,
        ])

    for r in range(2, ws.max_row + 1):
        has_flag = bool(ws.cell(row=r, column=9).value)
        for c in ws[r]:
            style_body_cell(c, fill=(FLAG_FILL if has_flag else None))

    set_column_widths(ws, [16, 38, 16, 14, 7, 12, 50, 28, 36])
    freeze_header_and_id_col(ws, header_row=1, id_col=2)


def write_building_blocks(ws, *, blocks_envelope: dict, apex_envelope: dict, spec: dict):
    blocks_data = blocks_envelope["data"]["blocks"]
    drone_score = apex_envelope["data"]["drone_score"]
    formula = apex_envelope["data"].get("formula_expression")

    ws.append(["block_id", "block_name", "weight_in_apex",
               "score", "contribution_to_drone_score",
               "gate_condition", "gate_triggered", "formula_expression"])
    style_header_row(ws, 1)

    contributions_data = apex_envelope["data"]["block_contributions"]
    for b in spec["building_blocks"]:
        if b["block_id"] == "BB_CAL_CONF":
            continue  # belongs on 5b sheet
        bid = b["block_id"]
        body = blocks_data.get(bid, {})
        contrib = contributions_data.get(bid, {}).get("contribution")
        ws.append([
            bid, b["block_name"], b["weight_in_drone_score_ppk"],
            body.get("score"),
            contrib,
            stringify(b.get("gate_condition")),
            "yes" if body.get("gate_triggered") else "no",
            stringify(b.get("formula_expression")),
        ])

    # APEX row
    ws.append([
        "drone_score", "APEX",
        sum(c.get("weight_in_ppk", 0) for c in contributions_data.values()),
        drone_score,
        drone_score,  # apex contributes to itself
        "(global gate — see Summary sheet)",
        "yes" if apex_envelope["data"].get("global_gate_triggered") else "no",
        formula,
    ])

    for r in range(2, ws.max_row):
        gate_on = ws.cell(row=r, column=7).value == "yes"
        for c in ws[r]:
            style_body_cell(c, fill=(GATE_FILL if gate_on else None))
    # apex row green
    for c in ws[ws.max_row]:
        style_body_cell(c, fill=APEX_FILL, bold=True)

    set_column_widths(ws, [18, 28, 16, 9, 22, 38, 14, 62])
    freeze_header_and_id_col(ws, header_row=1, id_col=2)


def write_cal_conf(ws, *, cal_envelope: dict):
    data = cal_envelope["data"]
    cal = data["cal_conf"]

    # banner row 1
    ws.append([
        "CAL_CONF — Calibration Confidence (NON-CONTRIBUTING parallel deliverable; "
        "weight_in_drone_score_ppk = 0; surfaced for operator visibility only)"
    ])
    ws.merge_cells("A1:G1")
    bcell = ws["A1"]
    bcell.font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=12)
    bcell.fill = CAL_CONF_BANNER_FILL
    bcell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    ws.row_dimensions[1].height = 28

    ws.append([])

    # block summary
    ws.append(["block_id", "block_name", "display_name", "weight_in_drone_score_ppk",
               "score", "gate_condition", "gate_triggered"])
    style_header_row(ws, ws.max_row)
    ws.append([
        cal.get("block_id"), cal.get("block_name"), cal.get("display_name"),
        cal.get("weight_in_drone_score_ppk"),
        cal.get("score"),
        stringify(cal.get("gate_condition")),
        "yes" if cal.get("gate_triggered") else "no",
    ])
    for c in ws[ws.max_row]:
        style_body_cell(c, fill=CAL_CONF_FILL, bold=True)
    ws.append([])

    # formula
    ws.append(["formula_expression", cal.get("formula_expression")])
    for c in ws[ws.max_row]:
        style_body_cell(c, fill=CAL_CONF_FILL)
    ws.append([])

    # indicator contributions
    ws.append(["Indicator contributions"])
    style_header_row(ws, ws.max_row)
    ws.append(["indicator_id", "weight", "indicator_score", "contribution"])
    style_header_row(ws, ws.max_row)
    contribs = cal.get("indicator_contributions", {}) or {}
    for iid in sorted(contribs.keys()):
        c = contribs[iid]
        ws.append([iid, c.get("weight"), c.get("indicator_score"), c.get("contribution")])
        for cell in ws[ws.max_row]:
            style_body_cell(cell, fill=CAL_CONF_FILL)
    ws.append([])

    # spec note for context
    if data.get("cal_conf_note"):
        ws.append(["Spec context note"])
        style_header_row(ws, ws.max_row)
        ws.append([data["cal_conf_note"]])
        for cell in ws[ws.max_row]:
            style_body_cell(cell, fill=CAL_CONF_FILL)
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=7)

    set_column_widths(ws, [18, 38, 30, 26, 18, 60, 16])
    freeze_header_and_id_col(ws, header_row=3, id_col=2)


def write_drone_score(ws, *, apex_envelope: dict, spec: dict):
    d = apex_envelope["data"]
    drone_score = d["drone_score"]
    gate_triggered = d.get("global_gate_triggered")
    img_cap_score = d["block_contributions"]["BB_IMG_CAPTURE"]["block_score"]

    # ---- headline strip ----
    ws.append([f"DRONE_SCORE = {drone_score}   (workflow: {d.get('workflow')})"])
    ws["A1"].font = Font(name=FONT_NAME, bold=True, size=16, color="1F3864")
    ws.merge_cells("A1:G1")
    ws.append([])

    # ---- formula + global gate ----
    ws.append(["formula_expression", d.get("formula_expression")])
    ws.append(["raw_weighted_sum",   d.get("raw_weighted_sum")])
    ws.append(["global_gate_triggered", str(gate_triggered)])
    ws.append(["global_gate_condition", d.get("global_gate_condition")])
    ws.append(["global_gate_action",    d.get("global_gate_action")])
    ws.append(["block_score_observed (BB_IMG_CAPTURE)", img_cap_score])
    ws.append([])

    # ---- block contributions ----
    ws.append(["Block contributions"])
    style_header_row(ws, ws.max_row)
    ws.append(["block_id", "block_name", "weight_in_apex", "block_score",
               "contribution", "block_gate_triggered", "notes"])
    style_header_row(ws, ws.max_row)
    for bid, c in d["block_contributions"].items():
        ws.append([
            bid, c.get("block_display_name"),
            c.get("weight_in_ppk"),
            c.get("block_score"),
            c.get("contribution"),
            "yes" if c.get("block_gate_triggered") else "no",
            stringify(c.get("notes"), 120),
        ])
        for cell in ws[ws.max_row]:
            style_body_cell(cell, fill=(GATE_FILL if c.get("block_gate_triggered") else None))
    ws.append([])

    # ---- CAL_CONF parallel callout ----
    cal_parallel = d.get("cal_conf_parallel", {})
    ws.append(["CAL_CONF parallel deliverable (NOT part of drone_score)"])
    style_header_row(ws, ws.max_row)
    ws.append(["score", cal_parallel.get("score"),
               "weight_in_drone_score_ppk",
               cal_parallel.get("weight_in_drone_score_ppk"),
               cal_parallel.get("note") or ""])
    for cell in ws[ws.max_row]:
        style_body_cell(cell, fill=CAL_CONF_FILL)
    ws.append([])

    # ---- all flags table ----
    ws.append(["All flags aggregated across stages"])
    style_header_row(ws, ws.max_row)
    ws.append(["flag_id", "flag_name", "severity", "_origin_stage",
               "raised_by", "context"])
    style_header_row(ws, ws.max_row)
    flags = d.get("all_flags_aggregated", []) or []
    if not flags:
        ws.append(["(no flags raised in this run)"])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=6)
        for cell in ws[ws.max_row]:
            style_body_cell(cell)
    for f in flags:
        ws.append([
            f.get("flag_id"), f.get("flag_name"), f.get("severity"),
            f.get("_origin_stage"),
            f.get("raised_by"),
            stringify(f.get("context"), 180),
        ])
        for cell in ws[ws.max_row]:
            style_body_cell(cell, fill=FLAG_FILL)
    ws.append([])

    # ---- cross-doc handoff candidates ----
    ws.append(["Cross-document handoff candidates (_handoff_crossdoc_candidates)"])
    style_header_row(ws, ws.max_row)
    crossdoc = d.get("_handoff_crossdoc_candidates", []) or []
    if not crossdoc:
        ws.append(["(none — apex envelope does not currently list any candidates "
                   "needing base station / GCP / processing bundles to evaluate)"])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=7)
        for cell in ws[ws.max_row]:
            style_body_cell(cell, fill=CROSSDOC_FILL)
    else:
        ws.append(["flag_id", "flag_name", "target_subsystem", "severity", "context"])
        style_header_row(ws, ws.max_row)
        for c in crossdoc:
            ws.append([
                c.get("flag_id"), c.get("flag_name"),
                c.get("target_subsystem"), c.get("severity"),
                stringify(c.get("context")),
            ])
            for cell in ws[ws.max_row]:
                style_body_cell(cell, fill=CROSSDOC_FILL)

    set_column_widths(ws, [36, 28, 18, 22, 22, 64, 30])
    freeze_header_and_id_col(ws, header_row=3, id_col=2)


def write_recommendations_placeholder(ws):
    ws.append(["8 — Recommendations (placeholder)"])
    ws["A1"].font = Font(name=FONT_NAME, bold=True, size=15, color="1F3864")
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 24
    ws.append([])
    ws.append(["Recommendations file not found at outputs/07_recommendations.json."])
    ws.append(["Generate it with:"])
    ws.append(["python3 scripts/compute_recommendations.py paths.json"])
    cmd_cell = ws.cell(row=ws.max_row, column=1)
    cmd_cell.font = Font(name="Menlo", size=11, color="1F3864")
    cmd_cell.fill = PatternFill("solid", fgColor="F2F2F2")
    ws.append(["…then re-run this exporter."])
    ws.append([])
    ws.append(["Why this exists"])
    style_header_row(ws, ws.max_row)
    ws.append(["The recommendations engine joins the pipeline's numerical indicator scores"])
    ws.append(["with the customer-facing library text (Drone_Recommendations/drone_indicator_library_v2_1.json)."])
    ws.append(["Per cbmi_chain_library_pattern.md: pipeline is the source of scores;"])
    ws.append(["library is the source of customer language; the engine is the join."])
    for r in range(3, ws.max_row + 1):
        for c in ws[r]:
            style_body_cell(c)
    set_column_widths(ws, [100])


def write_recommendations(ws, rec: dict):
    decision = rec.get("decision", "unknown")
    decision_display = decision.upper().replace("_", " ")
    tier  = rec.get("tier_interpretation")
    apex  = rec.get("apex_score")
    rat   = rec.get("decision_rationale")
    lib_v = rec.get("library_version")
    sp_v  = rec.get("spec_version")
    note  = rec.get("_spec_version_note")
    summary = rec.get("summary", {})
    caveats = rec.get("_caveats", [])

    # ----- title row -----
    ws.append([f"8 — Recommendations for Drone PPK   ({rec.get('subsystem','drone_ppk')})"])
    ws["A1"].font = Font(name=FONT_NAME, bold=True, size=15, color="FFFFFF")
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("A1:I1")
    ws.row_dimensions[1].height = 28
    ws.append([])

    # ----- headline KV block (rows 3-9) -----
    ws.append(["Decision", decision_display])
    dec_row = ws.max_row
    ws.merge_cells(start_row=dec_row, start_column=2, end_row=dec_row, end_column=4)
    dec_cell = ws.cell(row=dec_row, column=2)
    dec_cell.fill = DECISION_FILLS.get(decision, DECISION_FILLS["unable_to_assess"])
    dec_cell.font = Font(name=FONT_NAME, bold=True, size=14,
                         color=DECISION_FG.get(decision, "FFFFFF"))
    dec_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=dec_row, column=1).font = BODY_BOLD
    ws.row_dimensions[dec_row].height = 30

    ws.append(["Apex score",       apex])
    ws.append(["Tier",              tier])
    ws.append(["Rationale",         rat])
    ws.append(["Library version",   lib_v])
    ws.append(["Pipeline spec",     sp_v])
    ws.append(["Spec/library note", note or "(spec and library versions tracked separately)"])

    for r in range(3, ws.max_row + 1):
        ws.cell(row=r, column=1).font = BODY_BOLD
        for c in ws[r]:
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.append([])

    # ----- summary counts table -----
    ws.append(["Indicator counts by level"])
    style_header_row(ws, ws.max_row)
    ws.append(["good", "minor", "review", "resurvey", "unknown", "hard_gates_fired"])
    style_header_row(ws, ws.max_row)
    counts_row = [
        summary.get("good_count", 0),
        summary.get("minor_count", 0),
        summary.get("review_count", 0),
        summary.get("resurvey_count", 0),
        summary.get("unknown_count", 0),
        ", ".join(summary.get("hard_gates_fired", [])) or "(none)",
    ]
    ws.append(counts_row)
    fill = LEVEL_FILLS["resurvey"] if summary.get("hard_gates_fired") else LEVEL_FILLS["good"]
    for c in ws[ws.max_row]:
        style_body_cell(c, fill=fill, bold=True)
    ws.append([])

    # ----- per-indicator detail table -----
    ws.append(["Per-indicator recommendations"])
    style_header_row(ws, ws.max_row)
    ws.append([
        "indicator_id", "name", "block", "score", "level",
        "band_label", "recommendation_text", "actions", "hard_gate_fired"
    ])
    style_header_row(ws, ws.max_row)

    for row in rec.get("indicators", []):
        band = row.get("matched_band") or {}
        lvl = band.get("level") or "unknown"
        text = row.get("verified_statement") if lvl == "good" else row.get("impact")
        actions = row.get("actions") or []
        actions_text = "\n".join(f"• {a}" for a in actions) if actions else ""
        hg = row.get("hard_gate_fired")
        ws.append([
            row.get("indicator_id"),
            row.get("name"),
            row.get("block"),
            row.get("score"),
            lvl,
            band.get("label") or "",
            stringify(text, 300),
            actions_text,
            "yes" if hg else "",
        ])
        rfill = LEVEL_FILLS.get(lvl, LEVEL_FILLS["unknown"])
        for c in ws[ws.max_row]:
            style_body_cell(c, fill=rfill)
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[ws.max_row].height = 56 if (text or actions_text) else 22
    ws.append([])

    # ----- caveats block -----
    ws.append(["Caveats"])
    style_header_row(ws, ws.max_row)
    if not caveats:
        ws.append(["No caveats — every pipeline score landed in a library band cleanly."])
        for c in ws[ws.max_row]:
            style_body_cell(c, fill=LEVEL_FILLS["good"])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=9)
    else:
        ws.append(["indicator_id", "code", "details"])
        style_header_row(ws, ws.max_row)
        for c in caveats:
            details = json.dumps({k: v for k, v in c.items() if k not in ("indicator_id", "code")}, default=str)
            ws.append([c.get("indicator_id"), c.get("code"), stringify(details, 200)])
            for cell in ws[ws.max_row]:
                style_body_cell(cell, fill=LEVEL_FILLS["resurvey"])
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    set_column_widths(ws, [22, 24, 22, 9, 11, 38, 60, 56, 14])
    # Freeze panes on A4 — locks title + Decision callout at top
    ws.freeze_panes = ws["A4"]


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def main() -> int:
    if len(sys.argv) != 2:
        print("usage: export_results_xlsx.py <paths.json>", file=sys.stderr)
        return 2

    cfg_path = Path(sys.argv[1]).resolve()
    root = cfg_path.parent
    config = json.loads(cfg_path.read_text())

    # Discover all input artifacts
    spec = json.loads((root / config["spec_file"]).read_text())
    inventory = json.loads((root / config["outputs"]["stage1_inventory"]).read_text())
    src_envelope = json.loads((root / config["outputs"]["stage2_source_fields"]).read_text())
    der_envelope = json.loads((root / config["outputs"]["stage3_derived"]).read_text())
    ind_envelope = json.loads((root / config["outputs"]["stage3_indicators"]).read_text())
    blk_envelope = json.loads((root / config["outputs"]["stage3_building_blocks"]).read_text())
    cal_envelope = json.loads((root / config["outputs"]["stage3_cal_conf"]).read_text())
    apex_envelope = json.loads((root / config["outputs"]["stage3_drone_score"]).read_text())

    # Load recommendations envelope if present (per RECOMMENDATIONS_BUILD_PROMPT STEP D)
    rec_path = root / "outputs" / "07_recommendations.json"
    rec = json.loads(rec_path.read_text()) if rec_path.exists() else None

    wb = Workbook()

    # Summary sheet
    summary_ws = wb.active
    summary_ws.title = "Summary"
    sheet_legend = [
        ["Summary",            "Headline drone_score, recommendation callout, global gate, flag rollup"],
        ["1_Source_Files",     "Spec-required source files vs Stage 1 inventory (survey-level found/not-found)"],
        ["2_Source_Fields",    f"All {spec['_meta']['audit_counts']['source_fields']} L1F_* source fields"],
        ["3_Derived_Fields",   f"All {spec['_meta']['audit_counts']['derived_fields']} L2D_* derived fields with notes"],
        ["4_Indicators",       f"All {spec['_meta']['audit_counts']['indicators']} L3I_* indicators with trace blocks"],
        ["5_Building_Blocks",  "The 3 blocks that feed drone_score (BB_IMG_CAPTURE / BB_ROVER_GNSS / BB_MISSION_EXEC)"],
        ["5b_CAL_CONF",        "CAL_CONF parallel deliverable — does NOT contribute to drone_score"],
        ["6_Drone_Score",      "Apex value, block contributions, all flags, cross-doc handoff candidates"],
        ["8_Recommendations",  "Per-indicator customer-facing recommendations joined from the indicator library"],
    ]
    write_summary(summary_ws, apex=apex_envelope["data"], spec=spec, blocks=blk_envelope,
                  cal_conf=cal_envelope["data"]["cal_conf"], sheet_legend=sheet_legend,
                  rec=rec)

    # 1_Source_Files
    write_source_files(wb.create_sheet("1_Source_Files"),
                       inventory=inventory["data"] if "data" in inventory else inventory,
                       spec=spec)

    # 2_Source_Fields
    write_source_fields(wb.create_sheet("2_Source_Fields"),
                        source_envelope=src_envelope, spec=spec)

    # 3_Derived_Fields
    write_derived_fields(wb.create_sheet("3_Derived_Fields"),
                         derived_envelope=der_envelope, spec=spec)

    # 4_Indicators
    write_indicators(wb.create_sheet("4_Indicators"),
                     indicators_envelope=ind_envelope, spec=spec)

    # 5_Building_Blocks (excluding CAL_CONF)
    write_building_blocks(wb.create_sheet("5_Building_Blocks"),
                          blocks_envelope=blk_envelope,
                          apex_envelope=apex_envelope, spec=spec)

    # 5b_CAL_CONF (parallel deliverable)
    write_cal_conf(wb.create_sheet("5b_CAL_CONF"), cal_envelope=cal_envelope)

    # 6_Drone_Score
    write_drone_score(wb.create_sheet("6_Drone_Score"),
                      apex_envelope=apex_envelope, spec=spec)

    # 8_Recommendations (or stub if 07 not present)
    rec_ws = wb.create_sheet("8_Recommendations")
    if rec is None:
        write_recommendations_placeholder(rec_ws)
    else:
        write_recommendations(rec_ws, rec)

    out_dir = root / "outputs"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "drone_results.xlsx"
    wb.save(out_path)

    # Path-relative print (guarded — /tmp may resolve to /private/tmp on macOS)
    try:
        rel = out_path.relative_to(root)
    except ValueError:
        rel = out_path
    apex_score = apex_envelope["data"]["drone_score"]
    print(f"drone_score = {apex_score}")
    if rec is not None:
        print(f"decision     = {rec.get('decision')}")
        print(f"tier         = {rec.get('tier_interpretation')}")
    else:
        print("decision     = (no 07_recommendations.json — 8_Recommendations sheet rendered as stub)")
    print(f"wrote {rel}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
