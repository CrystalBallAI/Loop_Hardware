#!/usr/bin/env python3
"""Export Base Station PPK pipeline results to an Excel workbook.

One sheet per provenance level (Summary + 6 stage levels). Base Station is a
SINGLE-RECORD subsystem (one base-station occupation per survey), so each
detail sheet has ONE value/score column, no per-point fan-out, no
aggregate column. There is NO 5b parallel deliverable.

All cells carry VALUES (not formulas) — the numbers are the authoritative
pipeline results and the file must display correctly with no recalc engine.

Usage:
    python3 scripts/export_results_xlsx.py paths.json
"""
from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Style constants — Arial throughout; bold white-on-dark header; light fills
# for N/A, gate-triggered cells, flag rows, crossdoc-candidate rows.
# ---------------------------------------------------------------------------

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")     # dark blue
TITLE_FILL = PatternFill("solid", fgColor="2F5496")       # mid blue
NA_FILL = PatternFill("solid", fgColor="F2F2F2")          # light gray
GATE_FILL = PatternFill("solid", fgColor="FFE699")        # light yellow
FLAG_CRIT_FILL = PatternFill("solid", fgColor="F4B084")   # peach (CRITICAL)
FLAG_HIGH_FILL = PatternFill("solid", fgColor="FFD9CC")   # light orange
FLAG_MED_FILL = PatternFill("solid", fgColor="FFF2CC")    # light yellow
HANDOFF_FILL = PatternFill("solid", fgColor="DDEBF7")     # light blue
TOTAL_FILL = PatternFill("solid", fgColor="E2EFDA")       # light green

HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=14)
SECTION_FONT = Font(name=FONT_NAME, bold=True, size=12)
LABEL_FONT = Font(name=FONT_NAME, bold=True, size=10)
BODY_FONT = Font(name=FONT_NAME, size=10)
TOTAL_FONT = Font(name=FONT_NAME, bold=True, size=11)
APEX_FONT = Font(name=FONT_NAME, bold=True, size=16, color="1F4E79")

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

SEVERITY_FILLS = {
    "CRITICAL": FLAG_CRIT_FILL,
    "HIGH": FLAG_HIGH_FILL,
    "MEDIUM": FLAG_MED_FILL,
    "LOW": FLAG_MED_FILL,
}

# Band-level color palette for the recommendations sheet.
LEVEL_FILLS = {
    "good":     TOTAL_FILL,        # light green
    "minor":    HANDOFF_FILL,      # light blue
    "review":   FLAG_MED_FILL,     # light yellow
    "resurvey": FLAG_CRIT_FILL,    # peach
    "unknown":  NA_FILL,
}

# Chain-decision callout colors (used on Summary + 8_Recommendations).
DECISION_FILLS = {
    "good_to_go":            PatternFill("solid", fgColor="70AD47"),  # green
    "review_recommended":    PatternFill("solid", fgColor="FFC000"),  # amber
    "resurvey_recommended":  PatternFill("solid", fgColor="C00000"),  # red
    "unable_to_assess":      PatternFill("solid", fgColor="808080"),  # gray
}
DECISION_FG = {
    "good_to_go":            "FFFFFF",
    "review_recommended":    "000000",
    "resurvey_recommended":  "FFFFFF",
    "unable_to_assess":      "FFFFFF",
}

MAX_STRING_LEN = 200


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _stringify(value: Any) -> Any:
    """Pass primitives through; dict/list → JSON string truncated to ~200 chars."""
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float, str)):
        return value
    try:
        s = json.dumps(value, default=str, sort_keys=True)
    except (TypeError, ValueError):
        s = str(value)
    if len(s) > MAX_STRING_LEN:
        s = s[: MAX_STRING_LEN - 3] + "..."
    return s


def _set_header(ws, row: int, headers: list[str]) -> None:
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        c.border = BORDER


def _set_widths(ws, widths: list[float]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_row(ws, row: int, values: list[Any], fill: PatternFill | None = None,
               font: Font | None = None) -> None:
    for col, v in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=_stringify(v))
        c.font = font or BODY_FONT
        if fill is not None:
            c.fill = fill
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)
        c.border = BORDER


def _kv(ws, row: int, label: str, value: Any, value_font: Font | None = None,
        value_fill: PatternFill | None = None) -> int:
    lc = ws.cell(row=row, column=1, value=label)
    lc.font = LABEL_FONT
    lc.alignment = Alignment(horizontal="left", vertical="top")
    vc = ws.cell(row=row, column=2, value=_stringify(value))
    vc.font = value_font or BODY_FONT
    vc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    if value_fill is not None:
        vc.fill = value_fill
    return row + 1


def _section(ws, row: int, title: str, span: int = 4) -> int:
    c = ws.cell(row=row, column=1, value=title)
    c.font = SECTION_FONT
    c.alignment = Alignment(horizontal="left", vertical="center")
    if span > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    return row + 1


def _severity_fill(sev: str | None) -> PatternFill | None:
    if not sev:
        return None
    return SEVERITY_FILLS.get(sev.upper())


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _build_summary(wb: Workbook, spec: dict, s2: dict, s3d: dict, env_3d: dict,
                   rec: dict | None = None) -> None:
    ws = wb.create_sheet("Summary")
    d = s3d["data"]
    meta = d["stage3d_meta"]
    gg = d["global_gate"]

    # Title row
    tc = ws.cell(row=1, column=1, value="Base Station Confidence Score — Results Summary")
    tc.font = TITLE_FONT
    tc.fill = TITLE_FILL
    tc.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws.row_dimensions[1].height = 22

    row = 3
    row = _section(ws, row, "Apex")
    row = _kv(ws, row, "Subsystem", "base_station_ppk")
    row = _kv(ws, row, "Spec version", spec["_meta"]["version"])
    row = _kv(ws, row, "Apex score id", meta.get("apex_score_id"))
    row = _kv(ws, row, "Apex display name", meta.get("apex_display_name"))
    row = _kv(ws, row, "Workflow", meta.get("workflow"))
    row = _kv(ws, row, "Phase", meta.get("phase"))

    # Apex score — big bold
    lc = ws.cell(row=row, column=1, value="Base Station Score")
    lc.font = LABEL_FONT
    vc = ws.cell(row=row, column=2, value=d["base_station_score"])
    vc.font = APEX_FONT
    vc.fill = TOTAL_FILL if not gg["triggered"] else FLAG_CRIT_FILL
    row += 1

    row = _kv(ws, row, "Apex formula", d["apex_formula_spec"])
    row = _kv(ws, row, "Generated at", env_3d.get("generated_at"))
    row += 1

    # ---- Recommendation callout (new) — only when 07_recommendations.json exists ----
    if rec is not None:
        row = _section(ws, row, "Recommendation (from library v2.1)")
        decision = rec.get("decision") or "unable_to_assess"
        # Big colored decision cell
        lc = ws.cell(row=row, column=1, value="Decision")
        lc.font = LABEL_FONT
        vc = ws.cell(row=row, column=2, value=decision.upper().replace("_", " "))
        vc.font = Font(name=FONT_NAME, bold=True, size=14, color=DECISION_FG.get(decision, "000000"))
        vc.fill = DECISION_FILLS.get(decision, NA_FILL)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 26
        row += 1
        row = _kv(ws, row, "Tier", rec.get("tier_interpretation"))
        row = _kv(ws, row, "Rationale", rec.get("decision_rationale"))
        s = rec.get("summary", {})
        row = _kv(ws, row, "Indicators by level",
                  f"good={s.get('good_count', 0)}  minor={s.get('minor_count', 0)}  "
                  f"review={s.get('review_count', 0)}  resurvey={s.get('resurvey_count', 0)}")
        if s.get("hard_gates_fired"):
            row = _kv(ws, row, "Hard gates fired", ", ".join(s["hard_gates_fired"]),
                      value_fill=FLAG_CRIT_FILL,
                      value_font=Font(name=FONT_NAME, bold=True))
        row += 1

    row = _section(ws, row, "Global gate")
    row = _kv(ws, row, "Triggered", gg["triggered"],
              value_fill=GATE_FILL if gg["triggered"] else None,
              value_font=Font(name=FONT_NAME, bold=True) if gg["triggered"] else None)
    row = _kv(ws, row, "Condition (spec)", gg["condition_spec"])
    row = _kv(ws, row, "Action (spec)", gg["action_spec"])
    row = _kv(ws, row, "Block score observed (BB_BASE_COMPLETE)", gg["block_score_observed"])
    row += 1

    # Flags by severity
    row = _section(ws, row, "Flags by severity")
    _set_header(ws, row, ["severity", "count"])
    row += 1
    for sev, count in sorted(d.get("flags_by_severity", {}).items()):
        _write_row(ws, row, [sev, count], fill=_severity_fill(sev))
        row += 1
    row += 1

    # Flags by origin stage
    row = _section(ws, row, "Flags by origin stage")
    _set_header(ws, row, ["origin stage", "count"])
    row += 1
    for stage, count in sorted(d.get("flags_by_origin_stage", {}).items()):
        _write_row(ws, row, [stage, count])
        row += 1
    row += 1

    # Total flags aggregated
    row = _kv(ws, row, "Total flags aggregated", meta.get("total_flags_aggregated"))
    row += 1

    # Handoff crossdoc candidates
    row = _section(ws, row, "Handoff crossdoc candidates (deferred — need rover/drone bundle)")
    _set_header(ws, row, ["flag_id", "flag_name", "severity", "covers_problems", "status"])
    row += 1
    for c in d.get("_handoff_crossdoc_candidates", []) or []:
        _write_row(ws, row, [
            c.get("flag_id"), c.get("flag_name"), c.get("severity"),
            c.get("covers_problems"), c.get("status"),
        ], fill=HANDOFF_FILL)
        row += 1
    row += 1

    # Legend
    row = _section(ws, row, "Sheet legend")
    _set_header(ws, row, ["sheet", "what it carries"])
    row += 1
    legend = [
        ("Summary", "headline apex score, gate state, flag rollups, crossdoc candidates"),
        ("1_Source_Files", "3 source-file types with survey-level found status + physical counts"),
        ("2_Source_Fields", "36 source fields (L1F_BASE_*) — one value column per field"),
        ("3_Derived_Fields", "24 derived fields (L2D_BASE_*) — value + kind + notes"),
        ("4_Indicators", "11 indicators (L3I_BASE_*) — score / band / gate / flags"),
        ("5_Building_Blocks", "3 blocks (BB_BASE_*) — weighted score, weight_in_apex, formula + apex row"),
        ("6_Base_Station_Score", "apex value, gate, contributions, flags table, crossdoc table"),
    ]
    for s_name, desc in legend:
        _write_row(ws, row, [s_name, desc])
        row += 1

    _set_widths(ws, [32, 70, 18, 18])
    ws.freeze_panes = "A2"


def _build_source_files(wb: Workbook, s1: dict, spec: dict) -> None:
    ws = wb.create_sheet("1_Source_Files")
    d = s1["data"]
    counts = d.get("counts", {})

    # Critical-set policy: only RINEX OBS is critical (operator can run with
    # OPLOG/FORM degraded). Encode this here for the role column.
    role_by_src = {
        "SRC_BASE_RINEX": "critical (RINEX OBS hard-fail)",
        "SRC_BASE_OPLOG": "optional (degrades integrity to unconfirmed)",
        "SRC_BASE_FORM":  "optional (gates downstream block to 0)",
    }
    physical_count_by_src = {
        "SRC_BASE_RINEX": f"obs={counts.get('rinex_obs', 0)}, nav={counts.get('rinex_nav', 0)} (files_total={counts.get('rinex_files_total', 0)})",
        "SRC_BASE_OPLOG": f"instance={counts.get('oplog_instance', 0)} (files_total={counts.get('oplog_files_total', 0)})",
        "SRC_BASE_FORM":  f"instance={counts.get('form_instance', 0)} (files_total={counts.get('form_files_total', 0)})",
    }

    headers = ["file_id", "file_name", "role", "found", "physical_count"]
    _set_header(ws, 1, headers)
    spec_by_id = {sf["file_id"]: sf for sf in spec["source_files"]}
    row = 2
    for sid in ("SRC_BASE_RINEX", "SRC_BASE_OPLOG", "SRC_BASE_FORM"):
        sf = spec_by_id.get(sid, {})
        entry = d["expected_source_files"].get(sid, {})
        found = entry.get("found")
        fill = TOTAL_FILL if found else FLAG_CRIT_FILL
        _write_row(ws, row, [
            sid,
            sf.get("file_name") or entry.get("file_name"),
            role_by_src.get(sid),
            bool(found),
            physical_count_by_src.get(sid),
        ], fill=fill)
        row += 1

    _set_widths(ws, [22, 32, 38, 10, 50])
    ws.freeze_panes = "B2"


def _build_source_fields(wb: Workbook, s2: dict, spec: dict) -> None:
    ws = wb.create_sheet("2_Source_Fields")
    headers = ["field_id", "field_name", "source (file_id)", "category", "value"]
    _set_header(ws, 1, headers)
    spec_by_id = {sf["field_id"]: sf for sf in spec["source_fields"]}
    row = 2
    for key in sorted(s2["data"]["source_fields"].keys()):
        # key shape: "L1F_BASE_NNN_<field_name>"
        parts = key.split("_", 3)
        if len(parts) >= 4:
            field_id = "_".join(parts[:3])
            field_name = parts[3]
        else:
            field_id = key
            field_name = ""
        sf_spec = spec_by_id.get(field_id, {})
        value = s2["data"]["source_fields"][key]
        _write_row(ws, row, [
            field_id,
            field_name,
            sf_spec.get("file_id"),
            sf_spec.get("category"),
            value,
        ])
        row += 1

    _set_widths(ws, [18, 38, 20, 18, 70])
    ws.freeze_panes = "B2"


def _build_derived_fields(wb: Workbook, s3a: dict, spec: dict) -> None:
    ws = wb.create_sheet("3_Derived_Fields")
    headers = ["derived_id", "derived_name", "kind", "value", "is_na", "notes"]
    _set_header(ws, 1, headers)
    spec_by_id = {df["derived_id"]: df for df in spec["derived_fields"]}
    row = 2
    for key in sorted(s3a["data"]["derived_fields"].keys()):
        parts = key.split("_", 3)
        if len(parts) >= 4:
            derived_id = "_".join(parts[:3])
            derived_name = parts[3]
        else:
            derived_id = key
            derived_name = ""
        d = s3a["data"]["derived_fields"][key]
        value = d.get("value")
        kind = d.get("kind")
        notes_list = d.get("_notes") or []
        notes_str = " | ".join(notes_list) if notes_list else ""
        is_na = value is None
        fill = NA_FILL if is_na else None
        _write_row(ws, row, [
            derived_id,
            derived_name,
            kind,
            value,
            is_na,
            notes_str,
        ], fill=fill)
        row += 1

    _set_widths(ws, [18, 32, 16, 60, 8, 80])
    ws.freeze_panes = "B2"


def _build_indicators(wb: Workbook, s3b: dict) -> None:
    ws = wb.create_sheet("4_Indicators")
    headers = [
        "indicator_id", "indicator_name", "building_block_id", "weight_in_block",
        "score", "band_matched", "gate_triggered", "gate_action_spec",
        "flags_raised",
    ]
    _set_header(ws, 1, headers)
    row = 2
    for key in sorted(s3b["data"]["indicator_traces"].keys()):
        t = s3b["data"]["indicator_traces"][key]
        gate_triggered = bool(t.get("gate_triggered"))
        fill = GATE_FILL if gate_triggered else None
        flags_str = ", ".join(t.get("flags_raised") or []) or ""
        _write_row(ws, row, [
            t.get("indicator_id"),
            t.get("indicator_name"),
            t.get("building_block_id"),
            t.get("weight_in_block"),
            t.get("score"),
            t.get("band_matched"),
            gate_triggered,
            t.get("gate_action_spec"),
            flags_str,
        ], fill=fill)
        row += 1

    _set_widths(ws, [18, 36, 20, 14, 8, 30, 14, 50, 32])
    ws.freeze_panes = "B2"


def _build_blocks(wb: Workbook, s3c: dict, s3d: dict) -> None:
    ws = wb.create_sheet("5_Building_Blocks")
    headers = [
        "block_id", "block_name", "display_name",
        "weight_in_apex", "score", "weighted_score_before_gate",
        "gate_triggered", "gate_triggered_by_indicator", "formula_spec",
    ]
    _set_header(ws, 1, headers)
    row = 2
    blocks = s3c["data"]["block_scores"]
    # Preserve spec-formula order (COMPLETE, SETUP, ENV) by reading from
    # apex contributions (which we already wrote in spec order at Stage 3d).
    ordered_ids = [c["block_id"] for c in s3d["data"]["contributions"]]
    seen: set[str] = set()
    for bid in ordered_ids:
        b = blocks.get(bid)
        if b is None:
            continue
        seen.add(bid)
        gate_triggered = bool(b.get("gate_triggered"))
        fill = GATE_FILL if gate_triggered else None
        _write_row(ws, row, [
            b.get("block_id"),
            b.get("block_name"),
            b.get("display_name"),
            b.get("weight_in_apex"),
            b.get("score"),
            b.get("weighted_score_before_gate"),
            gate_triggered,
            b.get("gate_triggered_by_indicator"),
            b.get("formula_spec"),
        ], fill=fill)
        row += 1
    # Any blocks not in contributions (shouldn't happen for base):
    for bid, b in blocks.items():
        if bid in seen:
            continue
        _write_row(ws, row, [
            b.get("block_id"), b.get("block_name"), b.get("display_name"),
            b.get("weight_in_apex"), b.get("score"),
            b.get("weighted_score_before_gate"), bool(b.get("gate_triggered")),
            b.get("gate_triggered_by_indicator"), b.get("formula_spec"),
        ])
        row += 1

    # APEX row
    apex_score = s3d["data"]["base_station_score"]
    _write_row(ws, row, [
        "APEX",
        "base_station_score",
        "Base Station Confidence Score",
        sum(c["weight_in_apex"] for c in s3d["data"]["contributions"]),  # 1.00
        apex_score,
        s3d["data"]["weighted_score_before_global_gate"],
        s3d["data"]["global_gate"]["triggered"],
        None,
        s3d["data"]["apex_formula_spec"],
    ], fill=TOTAL_FILL, font=TOTAL_FONT)
    row += 1

    _set_widths(ws, [22, 36, 36, 14, 10, 22, 14, 28, 70])
    ws.freeze_panes = "B2"


def _build_apex(wb: Workbook, s3d: dict) -> None:
    ws = wb.create_sheet("6_Base_Station_Score")
    d = s3d["data"]
    gg = d["global_gate"]

    # Title
    tc = ws.cell(row=1, column=1, value="Base Station Score (apex)")
    tc.font = TITLE_FONT
    tc.fill = TITLE_FILL
    tc.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    ws.row_dimensions[1].height = 22

    row = 3
    row = _section(ws, row, "Apex headline", span=7)
    row = _kv(ws, row, "Apex score id", d["stage3d_meta"].get("apex_score_id"))
    row = _kv(ws, row, "Apex display name", d["stage3d_meta"].get("apex_display_name"))

    # apex score
    lc = ws.cell(row=row, column=1, value="Base Station Score")
    lc.font = LABEL_FONT
    vc = ws.cell(row=row, column=2, value=d["base_station_score"])
    vc.font = APEX_FONT
    vc.fill = TOTAL_FILL if not gg["triggered"] else FLAG_CRIT_FILL
    row += 1

    row = _kv(ws, row, "Weighted score before global gate", d["weighted_score_before_global_gate"])
    row = _kv(ws, row, "Apex formula (spec)", d["apex_formula_spec"])
    row += 1

    # Global gate
    row = _section(ws, row, "Global gate", span=7)
    row = _kv(ws, row, "Triggered", gg["triggered"],
              value_fill=GATE_FILL if gg["triggered"] else None,
              value_font=Font(name=FONT_NAME, bold=True) if gg["triggered"] else None)
    row = _kv(ws, row, "Condition (spec)", gg["condition_spec"])
    row = _kv(ws, row, "Action (spec)", gg["action_spec"])
    row = _kv(ws, row, "Block score observed (BB_BASE_COMPLETE)", gg["block_score_observed"])
    row += 1

    # Contributions table
    row = _section(ws, row, "Block contributions (apex weighted sum)", span=7)
    _set_header(ws, row, ["block_id", "block_name", "weight_in_apex", "block_score", "contribution"])
    row += 1
    contrib_total = 0.0
    for c in d["contributions"]:
        _write_row(ws, row, [
            c.get("block_id"),
            c.get("block_name"),
            c.get("weight_in_apex"),
            c.get("block_score"),
            c.get("contribution"),
        ])
        try:
            contrib_total += float(c.get("contribution", 0.0))
        except (TypeError, ValueError):
            pass
        row += 1
    _write_row(ws, row, ["TOTAL", "(sum of contributions)", None, None, round(contrib_total, 1)],
               fill=TOTAL_FILL, font=TOTAL_FONT)
    row += 2

    # Flags table — all_flags_aggregated
    row = _section(ws, row, "Flags aggregated (all_flags_aggregated)", span=7)
    _set_header(ws, row, [
        "flag_id", "flag_name", "severity", "raised_at_stage_spec",
        "_origin_stage", "_origin_derived_field", "condition_value",
    ])
    row += 1
    for f in d.get("all_flags_aggregated", []) or []:
        sev = f.get("severity")
        _write_row(ws, row, [
            f.get("flag_id"),
            f.get("flag_name"),
            sev,
            f.get("raised_at_stage_spec"),
            f.get("_origin_stage"),
            f.get("_origin_derived_field"),
            f.get("condition_value"),
        ], fill=_severity_fill(sev))
        row += 1
    row += 1

    # Handoff crossdoc candidates table
    row = _section(ws, row, "Handoff crossdoc candidates (deferred — need rover/drone bundle)", span=7)
    _set_header(ws, row, [
        "flag_id", "flag_name", "severity", "covers_problems", "status", "note",
    ])
    row += 1
    for c in d.get("_handoff_crossdoc_candidates", []) or []:
        _write_row(ws, row, [
            c.get("flag_id"),
            c.get("flag_name"),
            c.get("severity"),
            c.get("covers_problems"),
            c.get("status"),
            c.get("note"),
        ], fill=HANDOFF_FILL)
        row += 1

    _set_widths(ws, [22, 36, 16, 26, 16, 36, 70])
    ws.freeze_panes = "A4"


def _build_recommendations(wb: Workbook, rec: dict) -> None:
    """8th sheet: customer-facing recommendations from Tier 2 library."""
    ws = wb.create_sheet("8_Recommendations")

    # Title
    tc = ws.cell(row=1, column=1, value="Recommendations (Tier 2 library v2.1)")
    tc.font = TITLE_FONT
    tc.fill = TITLE_FILL
    tc.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    ws.row_dimensions[1].height = 22

    # ---- Headline block ----
    row = 3
    decision = rec.get("decision") or "unable_to_assess"

    lc = ws.cell(row=row, column=1, value="Decision")
    lc.font = LABEL_FONT
    dc = ws.cell(row=row, column=2, value=decision.upper().replace("_", " "))
    dc.font = Font(name=FONT_NAME, bold=True, size=16, color=DECISION_FG.get(decision, "000000"))
    dc.fill = DECISION_FILLS.get(decision, NA_FILL)
    dc.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    ws.row_dimensions[row].height = 30
    row += 1

    row = _kv(ws, row, "Apex score",          rec.get("apex_score"))
    row = _kv(ws, row, "Tier interpretation", rec.get("tier_interpretation"))
    row = _kv(ws, row, "Rationale",           rec.get("decision_rationale"))
    row = _kv(ws, row, "Library version",     rec.get("library_version"))
    row = _kv(ws, row, "Pipeline spec",       rec.get("spec_version"))
    note = rec.get("_spec_version_note")
    if note:
        row = _kv(ws, row, "Spec/library note", note)
    row += 1

    # ---- Summary counts ----
    row = _section(ws, row, "Summary by indicator-band level")
    _set_header(ws, row, ["good", "minor", "review", "resurvey", "unknown", "hard_gates_fired"])
    row += 1
    s = rec.get("summary", {})
    _write_row(ws, row, [
        s.get("good_count", 0),
        s.get("minor_count", 0),
        s.get("review_count", 0),
        s.get("resurvey_count", 0),
        s.get("unknown_count", 0),
        ", ".join(s.get("hard_gates_fired") or []) or "—",
    ], fill=(FLAG_CRIT_FILL if s.get("hard_gates_fired") else TOTAL_FILL),
       font=TOTAL_FONT)
    row += 2

    # ---- Per-indicator detail ----
    row = _section(ws, row, "Per-indicator recommendations (every indicator, in library order)")
    _set_header(ws, row, [
        "indicator_id", "name", "block", "score", "level", "band_label",
        "recommendation_text", "actions", "hard_gate_fired",
    ])
    row += 1

    for ind in rec.get("indicators", []):
        band = ind.get("matched_band") or {}
        level = band.get("level") or "unknown"
        is_good = level == "good"
        text = ind.get("verified_statement") if is_good else (ind.get("impact") or "")
        actions = ind.get("actions") or []
        actions_str = "\n• " + "\n• ".join(actions) if actions else ""
        if actions_str:
            actions_str = actions_str.lstrip("\n")
        fill = LEVEL_FILLS.get(level, NA_FILL)

        for col, value in enumerate([
            ind.get("indicator_id"),
            ind.get("name"),
            ind.get("block"),
            ind.get("score"),
            level,
            band.get("label"),
            text,
            actions_str,
            bool(ind.get("hard_gate_fired")),
        ], 1):
            c = ws.cell(row=row, column=col, value=_stringify(value))
            c.font = BODY_FONT
            c.fill = fill
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            c.border = BORDER
        # Bump the row height a bit so wrapped impact / actions text is readable.
        ws.row_dimensions[row].height = 56 if (text and not is_good) or actions_str else 22
        row += 1
    row += 1

    # ---- Caveats (if any) ----
    caveats = rec.get("_caveats") or []
    row = _section(ws, row, f"Caveats ({len(caveats)})")
    if caveats:
        _set_header(ws, row, ["indicator_id", "code", "detail"])
        row += 1
        for cav in caveats:
            _write_row(ws, row, [
                cav.get("indicator_id"),
                cav.get("code"),
                _stringify({k: v for k, v in cav.items() if k not in ("indicator_id", "code")}),
            ], fill=FLAG_HIGH_FILL)
            row += 1
    else:
        _write_row(ws, row, ["No caveats — every pipeline score landed in a library band cleanly."])
        row += 1

    _set_widths(ws, [22, 24, 22, 9, 11, 38, 60, 56, 14])
    ws.freeze_panes = "A4"


def _build_recommendations_placeholder(wb: Workbook) -> None:
    """Stub sheet when outputs/07_recommendations.json is missing."""
    ws = wb.create_sheet("8_Recommendations")
    tc = ws.cell(row=1, column=1, value="Recommendations (Tier 2 library v2.1)")
    tc.font = TITLE_FONT
    tc.fill = TITLE_FILL
    tc.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws.row_dimensions[1].height = 22

    nc = ws.cell(row=3, column=1,
                 value="Recommendations file not found. Generate it with:")
    nc.font = LABEL_FONT
    cc = ws.cell(row=4, column=1, value="python3 scripts/compute_recommendations.py paths.json")
    cc.font = Font(name="Courier New", size=11, bold=True)
    cc.fill = NA_FILL
    cc2 = ws.cell(row=6, column=1,
                  value="…then re-run this exporter to populate the 8_Recommendations sheet.")
    cc2.font = BODY_FONT

    _set_widths(ws, [90])


# ---------------------------------------------------------------------------
# Build orchestration
# ---------------------------------------------------------------------------

def build_workbook(spec: dict, s1: dict, s2: dict, s3a: dict, s3b: dict,
                   s3c: dict, s3d_env: dict, rec: dict | None = None) -> Workbook:
    """Construct the full workbook from in-memory envelopes.

    If `rec` (recommendations) is provided, an 8_Recommendations sheet is
    appended and the Summary sheet carries a decision callout.
    """
    wb = Workbook()
    # Remove the default "Sheet"
    default = wb.active
    if default is not None:
        wb.remove(default)

    _build_summary(wb, spec, s2, s3d_env, s3d_env, rec=rec)
    _build_source_files(wb, s1, spec)
    _build_source_fields(wb, s2, spec)
    _build_derived_fields(wb, s3a, spec)
    _build_indicators(wb, s3b)
    _build_blocks(wb, s3c, s3d_env)
    _build_apex(wb, s3d_env)

    if rec is not None:
        _build_recommendations(wb, rec)
    else:
        _build_recommendations_placeholder(wb)
    return wb


def _load(root: Path, rel: str) -> dict:
    return json.loads((root / rel).read_text(encoding="utf-8"))


def main(argv: list[str]) -> int:
    if len(argv) != 2:
        print("usage: export_results_xlsx.py <paths.json>", file=sys.stderr)
        return 2

    config_path = Path(argv[1]).resolve()
    root = config_path.parent
    config = json.loads(config_path.read_text(encoding="utf-8"))

    spec = _load(root, config["spec_file"])
    s1 = _load(root, config["outputs"]["stage1_inventory"])
    s2 = _load(root, config["outputs"]["stage2_source_fields"])
    s3a = _load(root, config["outputs"]["stage3_derived"])
    s3b = _load(root, config["outputs"]["stage3_indicators"])
    s3c = _load(root, config["outputs"]["stage3_building_blocks"])
    s3d = _load(root, config["outputs"]["stage3_base_score"])

    # Optional: recommendations file (07) — produced by compute_recommendations.py.
    # If absent, exporter still works and writes a stub stating how to generate it.
    rec_path = root / "outputs" / "07_recommendations.json"
    rec = json.loads(rec_path.read_text(encoding="utf-8")) if rec_path.exists() else None

    wb = build_workbook(spec, s1, s2, s3a, s3b, s3c, s3d, rec=rec)

    out_path = root / "outputs" / "base_station_results.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    # Headline + path; guard relative_to so paths outside the project root
    # (e.g. /tmp test runs on macOS where /tmp → /private/tmp) don't crash
    # after a successful save.
    apex = s3d["data"]["base_station_score"]
    gg_triggered = s3d["data"]["global_gate"]["triggered"]
    try:
        rel = out_path.relative_to(root)
    except ValueError:
        rel = out_path
    print(f"base_station_score = {apex}  (global_gate_triggered = {gg_triggered})")
    if rec is not None:
        print(f"decision           = {rec.get('decision')}  "
              f"(tier = {rec.get('tier_interpretation')})")
    else:
        print("decision           = (07_recommendations.json missing — stub sheet emitted)")
    print(f"wrote {rel}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
