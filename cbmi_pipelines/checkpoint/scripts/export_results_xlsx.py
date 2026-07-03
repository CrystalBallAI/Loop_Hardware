#!/usr/bin/env python3
"""export_results_xlsx.py - flatten the Check Point (RTK) pipeline outputs (01..06 JSON)
into a single Excel workbook, one sheet per provenance level, with the computed values.

Sheets: Summary | 1_Source_Files | 2_Source_Fields | 3_Derived_Fields | 4_Indicators |
        5_Building_Blocks | 6_Check_Point_Score

Check Point is PER-POINT / MULTI-OCCUPATION, so source fields / derived / indicators get
ONE COLUMN PER CHECK_POINT-role occupation (point_id headers). Building blocks show
per-point columns PLUS the cross-point aggregate (mean - k*(100-min)). Scoring is
CHECK_POINT-role ONLY: GCP-role points are excluded from aggregation (owned by gcp_score)
and surfaced as an excluded list, not as columns. There is NO 05b parallel deliverable.

The apex check_point_score can be null (zero CHECK_POINT-role points -> FLG_CP_002) or 0
(global gate: every CHECK_POINT-role point killed -> FLG_CP_001); both are read verbatim
from 06.data.null_handling + 06.data.global_gate.

VALUES ONLY: every number is read verbatim from outputs/*.json (the authoritative pipeline
results). No Excel formulas -> displays correctly with no recalc engine and zero formula
errors. Each sheet names its source artifact.

Run: python3 scripts/export_results_xlsx.py paths.json [out.xlsx]
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT, bold=True, size=13, color="1F4E78")
SUB_FONT = Font(name=FONT, italic=True, size=9, color="808080")
BASE_FONT = Font(name=FONT, size=10)
KEY_FONT = Font(name=FONT, bold=True, size=10)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NA_FILL = PatternFill("solid", fgColor="F2F2F2")        # N/A rows
FLAG_FILL = PatternFill("solid", fgColor="FCE4D6")      # flag / gate cells
GATE_FILL = PatternFill("solid", fgColor="FFF2CC")      # gate-triggered cells
GCP_FILL = PatternFill("solid", fgColor="EDEDED")       # excluded GCP-role columns
AGG_FILL = PatternFill("solid", fgColor="E2EFDA")       # cross-point aggregate column
HANDOFF_FILL = PatternFill("solid", fgColor="FFF2CC")  # handoff / placeholder rows
# Recommendations palettes (8_Recommendations + Summary callout). CheckPoint band
# levels: good / minor / review / critical (critical == resurvey tier) + n/a.
LEVEL_FILLS = {
    "good": PatternFill("solid", fgColor="C6E0B4"), "minor": PatternFill("solid", fgColor="DDEBF7"),
    "review": PatternFill("solid", fgColor="FFF2CC"), "critical": PatternFill("solid", fgColor="F8CBAD"),
    "resurvey": PatternFill("solid", fgColor="F8CBAD"), "n/a": PatternFill("solid", fgColor="F2F2F2"),
    "unknown": PatternFill("solid", fgColor="F2F2F2"),
}
DECISION_FILLS = {
    "good_to_go": PatternFill("solid", fgColor="70AD47"), "review_recommended": PatternFill("solid", fgColor="FFC000"),
    "resurvey_recommended": PatternFill("solid", fgColor="C00000"), "unable_to_assess": PatternFill("solid", fgColor="808080"),
}
DECISION_FG = {"good_to_go": "FFFFFF", "review_recommended": "000000",
               "resurvey_recommended": "FFFFFF", "unable_to_assess": "FFFFFF"}

# Inventory key (per point) -> spec source-file id + role. RTK export is the only
# CRITICAL input (Stage 1 critical_set_policy); oplog/form degrade gracefully.
SRC_TYPES = [
    ("rtk_export", "SRC_CP_RTK_EXPORT", "CRITICAL"),
    ("oplog", "SRC_CP_OPLOG", "OPTIONAL (device-type-aware)"),
    ("form", "SRC_CP_FORM", "OPTIONAL (degrade-to-unconfirmed)"),
]


def _fmt(v, maxlen=200):
    if v is None:
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, (dict, list)):
        s = json.dumps(v, sort_keys=True)
        return s if len(s) <= maxlen else s[: maxlen - 3] + "..."
    return v


def _write_table(ws, title, source, headers, rows, freeze_col=1):
    """Generic table: title + source + dark header row + bordered body. Freezes the
    header row AND the first `freeze_col` label columns so point columns scroll under
    fixed labels."""
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Source: {source}"
    ws["A2"].font = SUB_FONT
    hr = 4
    for c, h in enumerate(headers, 1):
        cell = ws.cell(hr, c, h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    for r, row in enumerate(rows, hr + 1):
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val)
            cell.font = BASE_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=isinstance(val, str) and len(str(val)) > 40)
    ws.freeze_panes = ws.cell(hr + 1, freeze_col + 1)
    for c in range(1, len(headers) + 1):
        width = len(str(headers[c - 1]))
        for row in rows:
            cellv = row[c - 1] if c - 1 < len(row) else None
            width = max(width, min(60, len(str(cellv)) if cellv is not None else 0))
        ws.column_dimensions[get_column_letter(c)].width = min(60, max(11, width + 2))
    return hr


def build(root: Path, config: dict, out_path: Path, rec: dict | None = None):
    spec = json.loads((root / config["spec_file"]).read_text())
    counts = spec["_meta"]["counts"]
    out = config["outputs"]
    env = {k: json.loads((root / out[v]).read_text())
           for k, v in {"inv": "stage1_inventory", "sf": "stage2_source_fields",
                        "der": "stage3_derived", "ind": "stage3_indicators",
                        "blk": "stage3_building_blocks",
                        "apex": "stage3_check_point_score"}.items()}
    O = {k: e["data"] for k, e in env.items()}
    apex_env = env["apex"]
    d = O["apex"]
    flag_name = {f["flag_id"]: f["flag_name"] for f in spec["flags"]}

    # ---- point partition: CHECK_POINT-role columns only; GCP-role excluded ----
    sf_points = O["sf"]["points"]
    cp_pids = [p["point_id"] for p in sf_points if p.get("device_role") == "CHECK_POINT"]
    gcp_pids = [p["point_id"] for p in sf_points if p.get("device_role") == "GCP"]
    role_by_pid = {p["point_id"]: p.get("device_role") for p in sf_points}
    pids = cp_pids  # the columns we fan out
    sf_by_pid = {p["point_id"]: p for p in sf_points}
    der_by_pid = {p["point_id"]: p for p in O["der"]["points"]}
    ind_by_pid = {p["point_id"]: p for p in O["ind"]["points"]}
    ppb_by_pid = {p["point_id"]: p for p in O["blk"]["per_point_blocks"]}

    apex_score = d["check_point_score"]
    apex_str = "null" if apex_score is None else apex_score
    eff_cp = d["stage3d_meta"]["effective_check_point_count"]

    wb = Workbook()
    wb.remove(wb.active)

    # ========================= Summary =========================
    ws = wb.create_sheet("Summary")
    ws["A1"] = "Check Point (RTK) - Confidence Score Results"
    ws["A1"].font = Font(name=FONT, bold=True, size=15, color="1F4E78")
    ws["A2"] = (f"spec {apex_env['spec_version']}  |  subsystem {config['subsystem']}  |  "
                f"survey {config['survey_id']}  |  generated {apex_env.get('generated_at', '')}")
    ws["A2"].font = SUB_FONT
    sd = d["stage3d_meta"]
    kv = [
        ("check_point_score (apex)", apex_str),
        ("effective_check_point_count", eff_cp),
        ("check_point_role_count", sd["check_point_role_count"]),
        ("excluded GCP-role points", ", ".join(gcp_pids) or "(none)"),
        ("apex formula", d["apex_formula_spec"]),
        ("weighted score before global gate", d["weighted_score_before_global_gate"]),
        ("global gate triggered", _fmt(d["global_gate"]["triggered"])),
        ("global gate condition", d["global_gate"]["condition_spec"]),
        ("null handling", d["null_handling"]["condition_spec"]),
        ("total flags aggregated", sd["total_flags_aggregated"]),
        ("flags by severity", _fmt(d["flags_by_severity"]) or "(none)"),
        ("flags by origin stage", _fmt(d["flags_by_origin_stage"]) or "(none)"),
        ("handoff crossdoc candidates", _fmt(d["_handoff_crossdoc_candidates"]) or "(none)"),
        ("apex weight-sum audit ok", _fmt(sd["apex_weight_sum_audit"]["ok"])),
    ]
    r = 4
    for k, v in kv:
        ws.cell(r, 1, k).font = KEY_FONT
        ws.cell(r, 2, v).font = BASE_FONT
        ws.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    # optional recommendation callout (D4)
    if rec:
        r = _summary_rec_callout(ws, r + 1, rec) + 1
    ws.cell(r + 1, 1, "Sheets (one per provenance level):").font = KEY_FONT
    legend = [
        ("1_Source_Files", f"{counts['source_files']} source-file types, per-point presence (Stage 1)"),
        ("2_Source_Fields", f"{counts['source_fields']} source fields, one column per point (Stage 2)"),
        ("3_Derived_Fields", f"{counts['derived_fields']} derived fields (15 per-point + 1 survey) (Stage 3a)"),
        ("4_Indicators", f"{counts['indicators']} indicators, per-point score/band/gate (Stage 3b)"),
        ("5_Building_Blocks", f"{counts['building_blocks']} blocks: per-point + cross-point aggregate + apex (Stage 3c/3d)"),
        ("6_Check_Point_Score", "apex score (or null), global gate, contributions, flags (Stage 3d)"),
        ("8_Recommendations", "per-point recommendations (placeholder until 07 is computed)"),
    ]
    for i, (s, desc) in enumerate(legend, r + 2):
        ws.cell(i, 1, s).font = Font(name=FONT, bold=True, size=10)
        ws.cell(i, 2, desc).font = BASE_FONT
    note_r = r + 2 + len(legend) + 1
    ws.cell(note_r, 1, "Layout note:").font = KEY_FONT
    ws.cell(note_r, 2, "Per-point fan-out: one column per CHECK_POINT-role occupation. "
                       "GCP-role points are excluded from scoring (owned by gcp_score). "
                       "No 05b: Check Point is a 3-block model with no parallel deliverable. "
                       "Values are verbatim pipeline results (no Excel formulas).").font = BASE_FONT
    ws.cell(note_r, 2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 95

    # ===================== 1_Source_Files =====================
    inv_points = {p["point_id"]: p for p in O["inv"]["points"]}
    inv_sum = O["inv"]["summary"]
    present_key = {"rtk_export": "points_with_export", "oplog": "points_with_oplog",
                   "form": "points_with_form"}
    headers = ["file_id", "spec_file_name", "inventory_key", "role"] + cp_pids + ["present_count"]
    rows = []
    spec_fname = {s["file_id"]: s["file_name"] for s in spec["source_files"]}
    for inv_key, file_id, role in SRC_TYPES:
        row = [file_id, spec_fname.get(file_id, ""), inv_key, role]
        for pid in cp_pids:
            entry = inv_points.get(pid, {}).get(inv_key)
            row.append(entry["filename"] if entry else "absent")
        row.append(inv_sum.get(present_key[inv_key], ""))
        rows.append(row)
    ws = wb.create_sheet("1_Source_Files")
    hr = _write_table(ws, f"Level 6 - Source Files ({counts['source_files']} types)",
                      out["stage1_inventory"], headers, rows, freeze_col=4)
    for r, row in enumerate(rows, hr + 1):
        for c, pid in enumerate(cp_pids, 5):
            if row[c - 1] == "absent":
                ws.cell(r, c).fill = NA_FILL

    # ===================== 2_Source_Fields =====================
    headers = ["field_id", "field_name", "source_file", "category"] + cp_pids
    rows = []
    for s in spec["source_fields"]:
        key = f"{s['field_id']}_{s['field_name']}"
        row = [s["field_id"], s["field_name"], s["file_id"], s.get("category", "")]
        for pid in cp_pids:
            row.append(_fmt(sf_by_pid[pid]["source_fields"].get(key)))
        rows.append(row)
    ws = wb.create_sheet("2_Source_Fields")
    _write_table(ws, f"Level 5 - Source Fields ({counts['source_fields']}), one column per point",
                 out["stage2_source_fields"], headers, rows, freeze_col=4)

    # ===================== 3_Derived_Fields =====================
    survey_keys = set(O["der"].get("survey_derived", {}).keys())
    headers = ["derived_id", "derived_name", "kind"] + cp_pids + ["survey_value", "is_na", "note"]
    rows = []
    na_row_idx = []
    for x in spec["derived_fields"]:
        key = f"{x['derived_id']}_{x['derived_name']}"
        row = [x["derived_id"], x["derived_name"], x.get("kind", "")]
        is_survey = key in survey_keys
        any_val = []
        if is_survey:
            row.extend(["" for _ in cp_pids])
            e = O["der"]["survey_derived"].get(key, {})
            row.append(_fmt(e.get("value")))
            any_val.append(e.get("value"))
            note = "; ".join(e.get("_notes", []))
        else:
            note = ""
            for pid in cp_pids:
                e = der_by_pid[pid]["derived_fields"].get(key, {})
                row.append(_fmt(e.get("value")))
                any_val.append(e.get("value"))
                if e.get("_notes") and not note:
                    note = "; ".join(e["_notes"])
            row.append("")  # survey_value blank for per-point fields
        is_na = all(v is None for v in any_val)
        row.append("TRUE" if is_na else "FALSE")
        row.append(_fmt(note, 300))
        rows.append(row)
        if is_na:
            na_row_idx.append(len(rows) - 1)
    ws = wb.create_sheet("3_Derived_Fields")
    hr = _write_table(ws, f"Level 4 - Derived Fields ({counts['derived_fields']}: "
                      f"{len(cp_pids) and counts['derived_fields'] - len(survey_keys)} per-point + "
                      f"{len(survey_keys)} survey)", out["stage3_derived"], headers, rows, freeze_col=3)
    ncols = len(headers)
    for idx in na_row_idx:
        for c in range(1, ncols + 1):
            ws.cell(hr + 1 + idx, c).fill = NA_FILL

    # ===================== 4_Indicators =====================
    # per point: score, band, gate, na  (4 cols/point) + leading 4 + trailing flags
    per_pt_cols = ["score", "band", "gate", "na"]
    headers = ["indicator_id", "indicator_name", "block_id", "weight_in_block"]
    for pid in cp_pids:
        headers += [f"{pid}|{c}" for c in per_pt_cols]
    headers += ["flags (union)"]
    rows = []
    gate_cells = []   # (row_idx, col_idx)
    na_rows = []
    for x in spec["indicators"]:
        iid = x["indicator_id"]
        # pull this indicator's trace from each point (traces keyed by id_name)
        traces = {}
        for pid in cp_pids:
            for tk, tv in ind_by_pid[pid]["indicator_traces"].items():
                if tv["indicator_id"] == iid:
                    traces[pid] = tv
                    break
        any_trace = next(iter(traces.values()), {})
        row = [iid, x["indicator_name"], any_trace.get("building_block_id", x.get("building_block_id")),
               any_trace.get("weight_in_block", x.get("weight_in_block"))]
        flag_union = set()
        row_na = True
        for pid in cp_pids:
            t = traces.get(pid, {})
            sc = "N/A" if t.get("na_redistribute") else t.get("score")
            row += [sc, t.get("band_matched"), _fmt(t.get("gate_triggered")), _fmt(t.get("na_redistribute"))]
            for fid in t.get("flags_raised", []):
                flag_union.add(flag_name.get(fid, fid))
            if not t.get("na_redistribute"):
                row_na = False
        row.append(", ".join(sorted(flag_union)))
        rows.append(row)
        ri = len(rows) - 1
        if row_na:
            na_rows.append(ri)
        # mark gate cells
        for j, pid in enumerate(cp_pids):
            gate_col = 4 + j * 4 + 3  # the 'gate' col for this point
            if traces.get(pid, {}).get("gate_triggered"):
                gate_cells.append((ri, gate_col))
    ws = wb.create_sheet("4_Indicators")
    hr = _write_table(ws, f"Level 3 - Indicators ({counts['indicators']}), per point",
                      out["stage3_indicators"], headers, rows, freeze_col=2)
    flags_col = len(headers)
    for ri, row in enumerate(rows):
        if row[flags_col - 1]:
            ws.cell(hr + 1 + ri, flags_col).fill = FLAG_FILL
    for ri, ci in gate_cells:
        ws.cell(hr + 1 + ri, ci).fill = GATE_FILL
    for ri in na_rows:
        for c in range(1, len(headers) + 1):
            ws.cell(hr + 1 + ri, c).fill = NA_FILL

    # ===================== 5_Building_Blocks =====================
    agg = O["blk"]["aggregated_blocks"]
    block_order = [c["block_id"] for c in d["contributions"]]  # spec-formula order
    headers = (["block_id", "block_name", "weight_in_apex"] + cp_pids
               + ["AGGREGATE (mean-k*(100-min))", "aggregator_k", "weight_sum_ok"])
    rows = []
    for bid in block_order:
        a = agg[bid]
        row = [bid, a["block_name"], a["weight_in_check_point_score"]]
        for pid in cp_pids:
            row.append(ppb_by_pid[pid]["block_scores"][bid]["score"])
        row += [a["aggregate_score"], a["aggregator_k"], _fmt(a["weight_sum_audit"]["ok"])]
        rows.append(row)
    ws = wb.create_sheet("5_Building_Blocks")
    hr = _write_table(ws, f"Level 1 - Building Blocks ({counts['building_blocks']}): "
                      f"per-point + cross-point aggregate", out["stage3_building_blocks"],
                      headers, rows, freeze_col=2)
    agg_col = 3 + len(cp_pids) + 1
    for r in range(hr + 1, hr + 1 + len(rows)):
        ws.cell(r, agg_col).fill = AGG_FILL
    # per-point per_point_score row + APEX row (read verbatim from 05/06; no formulas)
    pr = hr + len(rows) + 2
    ws.cell(pr, 1, "per_point_score (0.45C+0.35S+0.20E per point)").font = KEY_FONT
    for c, pid in enumerate(cp_pids, 4):
        ws.cell(pr, c, ppb_by_pid[pid]["per_point_score"]).font = KEY_FONT
    ar = pr + 1
    ws.cell(ar, 1, "APEX check_point_score = SUM(weight x aggregate)").font = KEY_FONT
    ws.cell(ar, agg_col, apex_str).font = KEY_FONT
    ws.cell(ar, agg_col).fill = AGG_FILL
    wr = ar + 1
    ws.cell(wr, 1, "apex weight sum (= 1.0)").font = SUB_FONT
    ws.cell(wr, 3, round(sum(d["apex_weights_used"].values()), 4)).font = SUB_FONT
    er = wr + 1
    ws.cell(er, 1, "effective_check_point_count").font = SUB_FONT
    ws.cell(er, 3, eff_cp).font = SUB_FONT

    # ===================== 6_Check_Point_Score =====================
    ws = wb.create_sheet("6_Check_Point_Score")
    ws["A1"] = "Level 0 - Check Point Score (apex)"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Source: {out['stage3_check_point_score']}"
    ws["A2"].font = SUB_FONT
    gg, nh = d["global_gate"], d["null_handling"]
    metrics = [
        ("check_point_score", apex_str),
        ("apex is null?", _fmt(apex_score is None)),
        ("effective_check_point_count", eff_cp),
        ("check_point_role_count", sd["check_point_role_count"]),
        ("excluded GCP-role points", ", ".join(gcp_pids) or "(none)"),
        ("weighted score before global gate", d["weighted_score_before_global_gate"]),
        ("global gate triggered", _fmt(gg["triggered"])),
        ("global gate condition", gg["condition_spec"]),
        ("global gate action", gg["action_spec"]),
        ("completeness_killed_by_point", _fmt(gg.get("completeness_killed_by_point"))),
        ("null handling", nh["condition_spec"]),
        ("no CHECK_POINT-role points?", _fmt(nh["no_check_point_role_points"])),
        ("apex formula", d["apex_formula_spec"]),
        ("apex weight-sum audit ok", _fmt(sd["apex_weight_sum_audit"]["ok"])),
        ("handoff crossdoc candidates", _fmt(d["_handoff_crossdoc_candidates"]) or "(none)"),
    ]
    rr = 4
    for name, val in metrics:
        ws.cell(rr, 1, name).font = KEY_FONT
        ws.cell(rr, 2, val).font = BASE_FONT
        ws.cell(rr, 2).alignment = Alignment(wrap_text=True, vertical="top")
        for c in (1, 2):
            ws.cell(rr, c).border = BORDER
        rr += 1
    # contributions table (spec-formula order)
    rr += 1
    ws.cell(rr, 1, "Block contributions (spec-formula order)").font = KEY_FONT
    rr += 1
    for c, h in enumerate(["block_id", "block_name", "weight_in_apex", "block_aggregate_score", "contribution"], 1):
        cell = ws.cell(rr, c, h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.border = BORDER
    rr += 1
    for con in d["contributions"]:
        for c, val in enumerate([con["block_id"], con["block_name"], con["weight_in_apex"],
                                 con["block_aggregate_score"], con["contribution"]], 1):
            cell = ws.cell(rr, c, val)
            cell.font = BASE_FONT
            cell.border = BORDER
        rr += 1
    # flags table (all_flags_aggregated, keeps _origin_stage + _origin_point)
    rr += 1
    ws.cell(rr, 1, f"Flags raised - {sd['total_flags_aggregated']} aggregated "
                   f"(severity {_fmt(d['flags_by_severity']) or 'none'}; "
                   f"by stage {_fmt(d['flags_by_origin_stage']) or 'none'})").font = KEY_FONT
    rr += 1
    for c, h in enumerate(["flag_id", "flag_name", "severity", "origin_stage", "origin_point", "condition_value"], 1):
        cell = ws.cell(rr, c, h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.border = BORDER
    rr += 1
    flag_rows = d["all_flags_aggregated"] or [{"flag_id": "(none)",
                "flag_name": "no flags raised - clean run", "severity": "",
                "_origin_stage": "", "_origin_point": "", "condition_value": ""}]
    for f in flag_rows:
        for c, val in enumerate([f.get("flag_id"), f.get("flag_name"), f.get("severity"),
                                 f.get("_origin_stage"), f.get("_origin_point"),
                                 _fmt(f.get("condition_value"))], 1):
            cell = ws.cell(rr, c, val)
            cell.font = BASE_FONT
            cell.fill = FLAG_FILL if d["all_flags_aggregated"] else NA_FILL
            cell.border = BORDER
        rr += 1
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 44
    for c in range(3, 7):
        ws.column_dimensions[get_column_letter(c)].width = 22

    # ===================== 8_Recommendations =====================
    ws = wb.create_sheet("8_Recommendations")
    if rec:
        write_recommendations(ws, rec)
    else:
        write_recommendations_placeholder(ws)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    got = {a: len(spec[a]) for a in ["source_files", "source_fields", "derived_fields",
                                     "indicators", "building_blocks"]}
    headline = {"check_point_score": apex_str, "effective_check_point_count": eff_cp,
                "cp_points": len(cp_pids), "gcp_excluded": len(gcp_pids)}
    return out_path, got, headline


def _summary_rec_callout(ws, r, rec):
    ws.cell(r, 1, "RECOMMENDATION").font = Font(name=FONT, bold=True, size=12, color="1F4E78")
    r += 1
    decision = rec.get("decision") or "unable_to_assess"
    ws.cell(r, 1, "decision").font = KEY_FONT
    dc = ws.cell(r, 2, decision.upper().replace("_", " "))
    dc.font = Font(name=FONT, bold=True, size=13, color=DECISION_FG.get(decision, "000000"))
    dc.fill = DECISION_FILLS.get(decision, NA_FILL)
    r += 1
    ss = rec.get("subsystem_summary", {})
    pts = (f"good {ss.get('good_points', 0)} / review {ss.get('review_points', 0)} / "
           f"resurvey {ss.get('resurvey_points', 0)}  (n={ss.get('n_points', 0)})")
    rows = [("tier", rec.get("tier_interpretation")), ("rationale", rec.get("decision_rationale")),
            ("points by decision", pts),
            ("effective_check_point_count", ss.get("effective_check_point_count")),
            ("library version", rec.get("library_version"))]
    if rec.get("_spec_version_note"):
        rows.append(("spec/library note", rec["_spec_version_note"]))
    for k, v in rows:
        ws.cell(r, 1, k).font = KEY_FONT
        ws.cell(r, 2, v).font = BASE_FONT
        ws.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    hg = ss.get("hard_gates_fired_by_point") or {}
    ws.cell(r, 1, "hard_gates_fired_by_point").font = KEY_FONT
    cc = ws.cell(r, 2, _fmt(hg) if hg else "(none — no critical-path indicators in this chain)")
    cc.font = BASE_FONT
    if hg:
        cc.fill = GATE_FILL
    r += 1
    return r


def write_recommendations_placeholder(ws):
    ws["A1"] = "8_Recommendations - PLACEHOLDER"
    ws["A1"].font = Font(name=FONT, bold=True, size=14, color="9C0006")
    ws["A1"].fill = HANDOFF_FILL
    ws["A3"] = "Recommendations have not been computed yet (outputs/07_recommendations.json not found)."
    ws["A3"].font = BASE_FONT
    ws["A5"] = "To populate this sheet (Deliverable 4):"
    ws["A5"].font = KEY_FONT
    ws["A6"] = "  python3 scripts/compute_recommendations.py paths.json"
    ws["A6"].font = Font(name="Courier New", size=10)
    ws["A7"] = "  python3 scripts/export_results_xlsx.py paths.json   (re-run to render)"
    ws["A7"].font = Font(name="Courier New", size=10)
    ws["A9"] = ("Once wired, this sheet shows: the chain decision callout, subsystem summary "
                "(good/review/resurvey point counts + hard_gates_fired_by_point), an indicator "
                "rollup, and a per-point fan-out of indicator recommendations.")
    ws["A9"].font = SUB_FONT
    ws["A9"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 110


def _hdr(ws, r, headers, start=1):
    for c, h in enumerate(headers, start):
        cell = ws.cell(r, c, h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def write_recommendations(ws, rec):
    """Full per-point recommendations sheet (D4 STEP D)."""
    NCOL = 9
    decision = rec.get("decision") or "unable_to_assess"
    tc = ws.cell(1, 1, "8_Recommendations - Check Point per-point (Tier 2 library)")
    tc.font = Font(name=FONT, bold=True, size=13, color="FFFFFF")
    tc.fill = HDR_FILL
    tc.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOL)
    ws.row_dimensions[1].height = 22

    rr = 3
    ws.cell(rr, 1, "Decision").font = KEY_FONT
    dc = ws.cell(rr, 2, decision.upper().replace("_", " "))
    dc.font = Font(name=FONT, bold=True, size=16, color=DECISION_FG.get(decision, "000000"))
    dc.fill = DECISION_FILLS.get(decision, NA_FILL)
    dc.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=4)
    ws.row_dimensions[rr].height = 30
    rr += 1

    def kv(label, val):
        nonlocal rr
        ws.cell(rr, 1, label).font = KEY_FONT
        c = ws.cell(rr, 2, _fmt(val) if isinstance(val, (dict, list)) else val)
        c.font = BASE_FONT
        c.alignment = Alignment(wrap_text=True, vertical="top")
        rr += 1

    kv("Apex score", "null" if rec.get("apex_score") is None else rec.get("apex_score"))
    kv("Tier interpretation", rec.get("tier_interpretation"))
    kv("Rationale", rec.get("decision_rationale"))
    kv("Library version", rec.get("library_version"))
    kv("Pipeline spec", rec.get("spec_version"))
    if rec.get("_spec_version_note"):
        kv("Spec/library note", rec["_spec_version_note"])
    kv("no_check_point_role_points", _fmt((rec.get("null_handling") or {}).get("no_check_point_role_points")))
    rr += 1

    # Subsystem summary
    ws.cell(rr, 1, "Subsystem summary").font = Font(name=FONT, bold=True, size=11, color="1F4E78")
    rr += 1
    _hdr(ws, rr, ["n_points", "good_points", "review_points", "resurvey_points",
                  "effective_check_point_count", "hard_gates_fired_by_point"])
    rr += 1
    ss = rec.get("subsystem_summary", {})
    hg = ss.get("hard_gates_fired_by_point") or {}
    for c, v in enumerate([ss.get("n_points", 0), ss.get("good_points", 0), ss.get("review_points", 0),
                           ss.get("resurvey_points", 0), ss.get("effective_check_point_count"),
                           _fmt(hg) if hg else "—"], 1):
        cell = ws.cell(rr, c, v)
        cell.font = BASE_FONT
        cell.border = BORDER
        if hg and c == 6:
            cell.fill = GATE_FILL
    rr += 2

    # Indicator rollup
    ws.cell(rr, 1, "Indicator rollup (worst across points)").font = Font(name=FONT, bold=True, size=11, color="1F4E78")
    rr += 1
    _hdr(ws, rr, ["indicator_id", "worst_score", "worst_level", "n_points_in_review",
                  "n_points_in_resurvey", "hard_gate_points"])
    rr += 1
    for iid in sorted(rec.get("indicator_rollup", {})):
        ro = rec["indicator_rollup"][iid]
        lvl = ro.get("worst_level_across_points", "good")
        vals = [iid, ro.get("worst_score_across_points"), lvl, ro.get("n_points_in_review", 0),
                ro.get("n_points_in_resurvey", 0), ", ".join(ro.get("hard_gate_points") or []) or "—"]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(rr, c, v)
            cell.font = BASE_FONT
            cell.border = BORDER
            cell.fill = LEVEL_FILLS.get(lvl, NA_FILL)
        rr += 1
    rr += 1

    # Per-point fan-out
    if not rec.get("points"):
        ws.cell(rr, 1, "No CHECK_POINT-role points — unable to assess (per-point detail empty).").font = SUB_FONT
        rr += 1
    for p in rec.get("points", []):
        hdr = ws.cell(rr, 1, f"Point {p['point_id']} ({p.get('device_role')}) — "
                             f"point_decision: {(p.get('point_decision') or '').upper()}")
        hdr.font = Font(name=FONT, bold=True, size=11, color="1F4E78")
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=NCOL)
        rr += 1
        _hdr(ws, rr, ["indicator_id", "name", "block", "score", "level", "band_label",
                      "recommendation_text", "actions", "hard_gate_fired"])
        rr += 1
        for ind in p.get("indicators", []):
            band = ind.get("matched_band") or {}
            level = band.get("level") or ("n/a" if ind.get("na_redistribute") else "unknown")
            is_good = level == "good"
            text = ind.get("verified_statement") if is_good else (ind.get("impact") or "")
            actions = ind.get("actions") or []
            actions_str = "• " + "\n• ".join(actions) if actions else ""
            fill = LEVEL_FILLS.get(level, NA_FILL)
            vals = [ind.get("indicator_id"), ind.get("name"), ind.get("block"), ind.get("score"),
                    level, band.get("label"), text, actions_str, _fmt(ind.get("hard_gate_fired"))]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(rr, c, _fmt(v) if isinstance(v, (dict, list)) else v)
                cell.font = BASE_FONT
                cell.fill = fill
                cell.border = BORDER
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.row_dimensions[rr].height = 56 if ((text and not is_good) or actions_str) else 22
            rr += 1
        rr += 1

    # Caveats
    caveats = rec.get("_caveats") or []
    ws.cell(rr, 1, f"Caveats ({len(caveats)})").font = Font(name=FONT, bold=True, size=11, color="1F4E78")
    rr += 1
    if caveats:
        _hdr(ws, rr, ["point_id", "indicator_id", "code", "detail"])
        rr += 1
        for cav in caveats:
            vals = [cav.get("point_id"), cav.get("indicator_id"), cav.get("code"),
                    _fmt({k: v for k, v in cav.items() if k not in ("point_id", "indicator_id", "code")})]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(rr, c, v)
                cell.font = BASE_FONT
                cell.border = BORDER
                cell.fill = FLAG_FILL
            rr += 1
    else:
        ws.cell(rr, 1, "No caveats — every (point, indicator) score landed in a library band cleanly.").font = BASE_FONT
        rr += 1

    for i, w in enumerate([22, 24, 22, 9, 11, 38, 60, 56, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"


def main(argv=None):
    ap = argparse.ArgumentParser(description="Export Check Point pipeline results to a single xlsx")
    ap.add_argument("config", nargs="?", default="paths.json")
    ap.add_argument("out", nargs="?", default="outputs/check_point_results.xlsx")
    args = ap.parse_args(argv)
    config_path = Path(args.config).resolve()
    config = json.loads(config_path.read_text())
    root = config_path.parent
    rec_path = root / "outputs" / "07_recommendations.json"
    rec = json.loads(rec_path.read_text()) if rec_path.exists() else None
    out, counts, headline = build(root, config, root / args.out, rec=rec)
    try:
        shown = out.relative_to(root)
    except ValueError:
        shown = out  # out path not under root (e.g. a /tmp test run); show absolute
    print(f"Wrote {shown}  (sheets: Summary + 6 levels + 8_Recommendations"
          f"{' [placeholder]' if rec is None else ''}, no 5b)")
    print(f"  spec counts: {counts}")
    print(f"  headline: check_point_score={headline['check_point_score']}  "
          f"effective_check_point_count={headline['effective_check_point_count']}  "
          f"(cp_points={headline['cp_points']}, gcp_excluded={headline['gcp_excluded']})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
