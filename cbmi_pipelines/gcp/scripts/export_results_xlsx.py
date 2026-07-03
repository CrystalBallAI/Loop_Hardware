#!/usr/bin/env python3
"""export_results_xlsx.py - flatten the GCP (PPK) pipeline outputs (01..06 JSON) into a
single Excel workbook, one sheet per provenance level, with the computed values.

Sheets: Summary | 1_Source_Files | 2_Source_Fields | 3_Derived_Fields | 4_Indicators |
        5_Building_Blocks | 6_GCP_Score | 8_Recommendations

GCP is PER-POINT / MULTI-OCCUPATION: source fields / derived / indicators fan out to ONE
COLUMN PER GCP-role occupation (point_id headers); building blocks show the cross-point
AGGREGATE rows (mean - k*(100-min)) PLUS a per-point block-scores table. Scoring is
GCP-role ONLY - CHECK_POINT-role points are excluded from aggregation (owned by
check_point_score) and surfaced as an excluded list. There is NO 5b parallel deliverable.

The apex gcp_score can be a number, 0 (global gate: every GCP-role point's coverage gate
fired), or null (null_handling: zero GCP-role points). All three render; the per-point
coverage-gate map and the null_handling block are surfaced on Summary + 6_GCP_Score.

8_Recommendations is a PLACEHOLDER until scripts/compute_recommendations.py emits
outputs/07_recommendations.json (Deliverable 4). When 07 exists, main() loads it.

VALUES ONLY: every number is read verbatim from outputs/*.json. No Excel formulas -> renders
with no recalc engine and zero formula errors.

Run: /opt/anaconda3/bin/python3 scripts/export_results_xlsx.py paths.json [out.xlsx]
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT, bold=True, size=13, color="1F4E78")
BIG_FONT = Font(name=FONT, bold=True, size=15, color="1F4E78")
SUB_FONT = Font(name=FONT, italic=True, size=9, color="808080")
BASE_FONT = Font(name=FONT, size=10)
KEY_FONT = Font(name=FONT, bold=True, size=10)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
# Palette (consistent with drone / base / check point)
GATE_FILL = PatternFill("solid", fgColor="F8CBAD")     # gate-fired rows / cells
FLAG_FILL = PatternFill("solid", fgColor="FCE4D6")     # flagged rows
NA_FILL = PatternFill("solid", fgColor="F2F2F2")       # N/A cells
APEX_FILL = PatternFill("solid", fgColor="C6E0B4")     # apex / sum rows
HANDOFF_FILL = PatternFill("solid", fgColor="FFF2CC")  # handoff candidate rows
CP_FILL = PatternFill("solid", fgColor="EDEDED")       # excluded CHECK_POINT-role rows
PT_TINTS = ["F4F9FE", "FFFBF4"]                         # alternating per-point column tint
# Recommendations palettes (8_Recommendations + Summary callout)
LEVEL_FILLS = {
    "good": PatternFill("solid", fgColor="C6E0B4"), "minor": PatternFill("solid", fgColor="DDEBF7"),
    "review": PatternFill("solid", fgColor="FFF2CC"), "resurvey": PatternFill("solid", fgColor="F8CBAD"),
    "unknown": PatternFill("solid", fgColor="F2F2F2"),
}
DECISION_FILLS = {
    "good_to_go": PatternFill("solid", fgColor="70AD47"), "review_recommended": PatternFill("solid", fgColor="FFC000"),
    "resurvey_recommended": PatternFill("solid", fgColor="C00000"), "unable_to_assess": PatternFill("solid", fgColor="808080"),
}
DECISION_FG = {"good_to_go": "FFFFFF", "review_recommended": "000000",
               "resurvey_recommended": "FFFFFF", "unable_to_assess": "FFFFFF"}

# Inventory slot -> spec source-file id + role + survey-level present-count key. GCP's 3 spec
# source files (RINEX/OPLOG/FORM) map onto 5 physical slots; RINEX OBS is the only critical input.
SRC_TYPES = [
    ("rinex_obs", "SRC_GCP_RINEX", "CRITICAL (only critical input)", "points_with_obs"),
    ("rinex_nav", "SRC_GCP_RINEX", "OPTIONAL (NAV: PDOP/acquisition)", "points_with_nav"),
    ("hardware", "SRC_GCP_RINEX", "OPTIONAL (4-tier header override)", "points_with_hardware"),
    ("oplog", "SRC_GCP_OPLOG", "OPTIONAL (DGPS device-type-aware)", "points_with_oplog"),
    ("form", "SRC_GCP_FORM", "OPTIONAL (degrade-to-unconfirmed)", "points_with_form"),
]
PT_SLOTS = ["form", "rinex_obs", "rinex_nav", "hardware", "oplog"]


def _fmt(v, maxlen=200):
    if v is None:
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, (dict, list)):
        s = json.dumps(v, sort_keys=True)
        return s if len(s) <= maxlen else s[: maxlen - 3] + "..."
    return v


def _tint(hexcolor):
    return PatternFill("solid", fgColor=hexcolor)


def _header_row(ws, r, headers, start=1, fill=HDR_FILL):
    for c, h in enumerate(headers, start):
        cell = ws.cell(r, c, h)
        cell.font = HDR_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def _autosize(ws, headers, rows, start=1, cap=60):
    for i, h in enumerate(headers):
        c = start + i
        width = len(str(h))
        for row in rows:
            v = row[i] if i < len(row) else None
            width = max(width, min(cap, len(str(v)) if v is not None else 0))
        ws.column_dimensions[get_column_letter(c)].width = min(cap, max(11, width + 2))


def build(root: Path, config: dict, out_path: Path, rec: dict | None = None):
    spec = json.loads((root / config["spec_file"]).read_text())
    counts = spec["_meta"]["counts"]          # GCP uses _meta.counts (NOT audit_counts)
    out = config["outputs"]
    env = {k: json.loads((root / out[v]).read_text())
           for k, v in {"inv": "stage1_inventory", "sf": "stage2_source_fields",
                        "der": "stage3_derived", "ind": "stage3_indicators",
                        "blk": "stage3_building_blocks", "apex": "stage3_gcp_score"}.items()}
    O = {k: e["data"] for k, e in env.items()}
    apex_env = env["apex"]
    d = O["apex"]
    flag_name = {f["flag_id"]: f["flag_name"] for f in spec["flags"]}

    # ---- point partition: GCP-role columns; CHECK_POINT excluded (owned by check_point_score)
    sf_points = O["sf"]["points"]
    gcp_pids = [p["point_id"] for p in sf_points if p.get("device_role") == "GCP"]
    cp_pids = [p["point_id"] for p in sf_points if p.get("device_role") == "CHECK_POINT"]
    sf_by_pid = {p["point_id"]: p for p in sf_points}
    der_by_pid = {p["point_id"]: p for p in O["der"].get("points", [])}
    ind_by_pid = {p["point_id"]: p for p in O["ind"].get("points", [])}
    ppb_by_pid = {p["point_id"]: p for p in O["blk"].get("per_point_blocks", [])}
    agg = O["blk"].get("aggregated_blocks", {})

    sd = d["stage3d_meta"]
    gcp_count = sd["gcp_role_point_count"]
    apex_score = d["gcp_score"]
    is_null = apex_score is None
    apex_str = "N/A (null_handling)" if is_null else apex_score
    gg, nh = d["global_gate"], d["null_handling"]
    weight_ok = sd["apex_weight_sum_audit"]["ok"]
    block_order = [c["block_id"] for c in d["contributions"]] or list(agg.keys())

    def pt_fill(j):
        return _tint(PT_TINTS[j % 2])

    wb = Workbook()
    wb.remove(wb.active)

    # ============================ Summary ============================
    ws = wb.create_sheet("Summary")
    ws["A1"] = "GCP (PPK) - Confidence Score Results"
    ws["A1"].font = BIG_FONT
    ws["A2"] = (f"spec {apex_env['spec_version']}  |  subsystem {config['subsystem']}  |  "
                f"survey {config['survey_id']}  |  generated {apex_env.get('generated_at', '')}")
    ws["A2"].font = SUB_FONT
    # headline
    ws["A4"] = "GCP SCORE"
    ws["A4"].font = KEY_FONT
    hcell = ws.cell(4, 2, "GCP SCORE = N/A (null_handling triggered)" if is_null else apex_score)
    hcell.font = Font(name=FONT, bold=True, size=14,
                      color="9C0006" if is_null else "1F4E78")
    hcell.fill = GATE_FILL if is_null else APEX_FILL
    r = 6
    kv = [
        ("apex formula", d["apex_formula_spec"]),
        ("gcp_role_point_count", gcp_count),
        ("excluded CHECK_POINT-role points", ", ".join(cp_pids) or "(none)"),
        ("weighted score before global gate", _na(d["weighted_score_before_global_gate"], is_null)),
        ("total flags aggregated", sd["total_flags_aggregated"]),
        ("flags by severity", _fmt(d["flags_by_severity"]) or "(none)"),
        ("flags by origin stage", _fmt(d["flags_by_origin_stage"]) or "(none)"),
        ("apex weight-sum audit ok", _fmt(weight_ok)),
    ]
    for k, v in kv:
        ws.cell(r, 1, k).font = KEY_FONT
        ws.cell(r, 2, v).font = BASE_FONT
        ws.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    # null_handling block (prominent - the per-point chain's distinctive feature)
    r += 1
    ws.cell(r, 1, "null_handling").font = Font(name=FONT, bold=True, size=11, color="9C0006")
    r += 1
    for k, v in [("condition_spec", nh["condition_spec"]),
                 ("no_gcp_role_points", _fmt(nh["no_gcp_role_points"]))]:
        ws.cell(r, 1, k).font = KEY_FONT
        cc = ws.cell(r, 2, v)
        cc.font = BASE_FONT
        cc.alignment = Alignment(wrap_text=True, vertical="top")
        if k == "no_gcp_role_points" and nh["no_gcp_role_points"]:
            cc.fill = GATE_FILL
        r += 1
    # global gate + per-point coverage map
    r += 1
    ws.cell(r, 1, "global_gate").font = Font(name=FONT, bold=True, size=11, color="1F4E78")
    r += 1
    for k, v in [("triggered", _fmt(gg["triggered"])), ("condition_spec", gg["condition_spec"]),
                 ("action_spec", gg["action_spec"])]:
        ws.cell(r, 1, k).font = KEY_FONT
        ws.cell(r, 2, v).font = BASE_FONT
        ws.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")
        if k == "triggered" and gg["triggered"]:
            ws.cell(r, 2).fill = GATE_FILL
        r += 1
    ws.cell(r, 1, "coverage_gate_fired_by_point").font = KEY_FONT
    r += 1
    for pid, fired in (gg.get("coverage_gate_fired_by_point") or {}).items():
        ws.cell(r, 1, f"   {pid}").font = BASE_FONT
        cc = ws.cell(r, 2, "GATE FIRED" if fired else "ok")
        cc.font = BASE_FONT
        if fired:
            cc.fill = GATE_FILL
        r += 1
    # handoff candidates (GCP does not emit this key)
    r += 1
    ws.cell(r, 1, "_handoff_crossdoc_candidates").font = KEY_FONT
    has_handoff = "_handoff_crossdoc_candidates" in d
    ws.cell(r, 2, _fmt(d.get("_handoff_crossdoc_candidates")) if has_handoff
            else "(not currently emitted by GCP)").font = BASE_FONT
    r += 1
    # optional recommendation callout
    if rec:
        r = _summary_rec_callout(ws, r + 1, rec)
    # legend
    r += 1
    ws.cell(r, 1, "Sheets (one per provenance level):").font = KEY_FONT
    legend = [
        ("1_Source_Files", f"{counts['source_files']} spec files / 5 slots + per-point Points block (Stage 1)"),
        ("2_Source_Fields", f"{counts['source_fields']} source fields, one column per point (Stage 2)"),
        ("3_Derived_Fields", f"{counts['derived_fields']} derived fields, all per-point (Stage 3a)"),
        ("4_Indicators", f"{counts['indicators']} indicators, per-point scores + cross-point rollup (Stage 3b)"),
        ("5_Building_Blocks", f"{counts['building_blocks']} blocks: aggregate rows + per-point table + apex (Stage 3c/3d)"),
        ("6_GCP_Score", "apex (or null), global gate map, contributions, flags (Stage 3d)"),
        ("8_Recommendations", "per-point recommendations (placeholder until 07 is computed)"),
    ]
    for s, desc in legend:
        r += 1
        ws.cell(r, 1, s).font = Font(name=FONT, bold=True, size=10)
        ws.cell(r, 2, desc).font = BASE_FONT
    r += 2
    ws.cell(r, 1, "Layout note:").font = KEY_FONT
    ws.cell(r, 2, "Per-point fan-out: one column per GCP-role occupation. CHECK_POINT-role points "
                  "are excluded from scoring (owned by check_point_score). No 5b: GCP is a 3-block "
                  "model. Values are verbatim pipeline results (no Excel formulas).").font = BASE_FONT
    ws.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 96

    # ========================= 1_Source_Files =========================
    ws = wb.create_sheet("1_Source_Files")
    ws["A1"] = f"Level 6 - Source Files ({counts['source_files']} spec files / 5 inventory slots)"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Source: {out['stage1_inventory']}"
    ws["A2"].font = SUB_FONT
    inv_points = {p["point_id"]: p for p in O["inv"].get("points", [])}
    inv_sum = O["inv"].get("summary", {})
    spec_fname = {s["file_id"]: s["file_name"] for s in spec["source_files"]}
    # Table A: source-file types (survey-level)
    headers = ["inventory_slot", "spec_file_id", "spec_file_name", "role",
               f"present_count (/{len(gcp_pids)})"]
    _header_row(ws, 4, headers)
    rA = 5
    for slot, fid, role, sumk in SRC_TYPES:
        vals = [slot, fid, spec_fname.get(fid, ""), role, inv_sum.get(sumk, "")]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(rA, c, v)
            cell.font = BASE_FONT
            cell.border = BORDER
        rA += 1
    _autosize(ws, headers, [[slot, fid, spec_fname.get(fid, ""), role, ""] for slot, fid, _, _ in
              [(s[0], s[1], 0, 0) for s in SRC_TYPES] for role in [s[2] for s in SRC_TYPES][:1]])
    ws.column_dimensions["C"].width = 34
    ws.column_dimensions["D"].width = 38
    # Table B: per-point Points block
    pb_r = rA + 2
    ws.cell(pb_r, 1, "Points (per-point folders + per-source presence)").font = KEY_FONT
    pb_r += 1
    pheaders = ["point_id", "device_role", "device_type", "point_folder"] + PT_SLOTS
    _header_row(ws, pb_r, pheaders)
    pr = pb_r + 1
    for p in sf_points:
        pid = p["point_id"]
        ip = inv_points.get(pid, {})
        row = [pid, p.get("device_role"), p.get("device_type"), ip.get("point_folder", "")]
        for slot in PT_SLOTS:
            entry = ip.get(slot)
            tag = " (PLACEHOLDER)" if isinstance(entry, dict) and entry.get("status") == "PLACEHOLDER" else ""
            row.append(f"{entry['filename']}{tag}" if isinstance(entry, dict) and entry.get("filename") else "absent")
        excluded = p.get("device_role") == "CHECK_POINT"
        for c, v in enumerate(row, 1):
            cell = ws.cell(pr, c, v)
            cell.font = BASE_FONT
            cell.border = BORDER
            if excluded:
                cell.fill = CP_FILL
            elif c > 4 and v == "absent":
                cell.fill = NA_FILL
        pr += 1
    ws.freeze_panes = "A5"

    # ========================= 2_Source_Fields =========================
    ws = wb.create_sheet("2_Source_Fields")
    ws["A1"] = f"Level 5 - Source Fields ({counts['source_fields']}), one column per GCP point"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Source: {out['stage2_source_fields']}"
    ws["A2"].font = SUB_FONT
    headers = ["field_id", "field_name", "source_file", "category"] + gcp_pids
    _header_row(ws, 4, headers)
    rr = 5
    rows_for_size = []
    for s in spec["source_fields"]:
        key = f"{s['field_id']}_{s['field_name']}"
        row = [s["field_id"], s["field_name"], s["file_id"], s.get("category", "")]
        for pid in gcp_pids:
            v = sf_by_pid[pid]["source_fields"].get(key)
            row.append("N/A" if v is None and key not in sf_by_pid[pid]["source_fields"] else _fmt(v))
        rows_for_size.append(row)
        for c, v in enumerate(row, 1):
            cell = ws.cell(rr, c, v)
            cell.font = BASE_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=isinstance(v, str) and len(v) > 40)
            if c > 4:
                cell.fill = pt_fill(c - 5)
                if v == "N/A":
                    cell.fill = NA_FILL
        rr += 1
    _autosize(ws, headers, rows_for_size)
    ws.freeze_panes = "E5"

    # ========================= 3_Derived_Fields =========================
    ws = wb.create_sheet("3_Derived_Fields")
    ws["A1"] = f"Level 4 - Derived Fields ({counts['derived_fields']}, all per-point)"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Source: {out['stage3_derived']}"
    ws["A2"].font = SUB_FONT
    headers = ["derived_id", "derived_name", "kind"] + gcp_pids + ["is_na", "note"]
    _header_row(ws, 4, headers)
    rr = 5
    rows_for_size = []
    for x in spec["derived_fields"]:
        key = f"{x['derived_id']}_{x['derived_name']}"
        row = [x["derived_id"], x["derived_name"], x.get("kind", "")]
        vals, note = [], ""
        for pid in gcp_pids:
            e = der_by_pid[pid]["derived_fields"].get(key, {})
            row.append(_fmt(e.get("value")))
            vals.append(e.get("value"))
            if e.get("_notes") and not note:
                note = "; ".join(e["_notes"])
        na = all(v is None for v in vals)
        row += ["TRUE" if na else "FALSE", _fmt(note, 300)]
        rows_for_size.append(row)
        for c, v in enumerate(row, 1):
            cell = ws.cell(rr, c, v)
            cell.font = BASE_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=isinstance(v, str) and len(str(v)) > 40)
            if 3 < c <= 3 + len(gcp_pids):
                cell.fill = pt_fill(c - 4)
        if na:
            for c in range(1, len(headers) + 1):
                ws.cell(rr, c).fill = NA_FILL
        rr += 1
    _autosize(ws, headers, rows_for_size)
    ws.freeze_panes = "D5"

    # ========================= 4_Indicators =========================
    ws = wb.create_sheet("4_Indicators")
    ws["A1"] = f"Level 3 - Indicators ({counts['indicators']}), per-point scores + cross-point rollup"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Source: {out['stage3_indicators']}  (per-indicator aggregate is block-level; see sheet 5)"
    ws["A2"].font = SUB_FONT
    lead = ["indicator_id", "indicator_name", "block_id", "weight_in_block"]
    tail = ["aggregator", "aggregate_score", "band_matched(agg)", "gate_triggered_any_point", "flags_raised_union"]
    headers = lead + gcp_pids + tail
    _header_row(ws, 4, headers)
    rr = 5
    rows_for_size = []
    for x in spec["indicators"]:
        iid = x["indicator_id"]
        traces = {}
        for pid in gcp_pids:
            for tv in ind_by_pid[pid]["indicator_traces"].values():
                if tv["indicator_id"] == iid:
                    traces[pid] = tv
                    break
        anyt = next(iter(traces.values()), {})
        row = [iid, x["indicator_name"], anyt.get("building_block_id", x.get("building_block_id")),
               anyt.get("weight_in_block", x.get("weight_in_block"))]
        gate_any = False
        flag_union = set()
        per_pt_gate = []
        for pid in gcp_pids:
            t = traces.get(pid, {})
            row.append(t.get("score"))
            g = bool(t.get("gate_triggered"))
            per_pt_gate.append(g)
            gate_any = gate_any or g
            for fid in t.get("flags_raised", []):
                flag_union.add(flag_name.get(fid, fid))
        row += ["block-level (sheet 5)", "N/A (block-level)", "N/A",
                _fmt(gate_any), ", ".join(sorted(flag_union))]
        rows_for_size.append(row)
        for c, v in enumerate(row, 1):
            cell = ws.cell(rr, c, v)
            cell.font = BASE_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=isinstance(v, str) and len(str(v)) > 40)
            if 4 < c <= 4 + len(gcp_pids):
                cell.fill = GATE_FILL if per_pt_gate[c - 5] else pt_fill(c - 5)
        if gate_any:
            for c in range(1, 5):
                ws.cell(rr, c).fill = GATE_FILL
        if flag_union:
            ws.cell(rr, len(headers)).fill = FLAG_FILL
        rr += 1
    _autosize(ws, headers, rows_for_size)
    ws.freeze_panes = "E5"

    # ========================= 5_Building_Blocks =========================
    ws = wb.create_sheet("5_Building_Blocks")
    ws["A1"] = f"Level 1 - Building Blocks ({counts['building_blocks']}): aggregate + per-point"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Source: {out['stage3_building_blocks']}"
    ws["A2"].font = SUB_FONT
    # Table A: aggregated blocks
    headers = ["block_id", "block_name", "weight_in_apex", "aggregator_spec", "aggregator_k",
               "aggregate_score", "has_internal_gate", "gate_condition_spec", "gate_triggered(any_point)"]
    _header_row(ws, 4, headers)
    rr = 5
    rows_for_size = []
    for bid in block_order:
        a = agg.get(bid, {})
        gate_any = any(ppb_by_pid[pid]["block_scores"].get(bid, {}).get("gate_triggered")
                       for pid in gcp_pids)
        row = [bid, a.get("block_name"), a.get("weight_in_gcp_score"), a.get("aggregator_spec"),
               a.get("aggregator_k"), a.get("aggregate_score"), _fmt(a.get("has_internal_gate")),
               a.get("gate_condition_spec"), _fmt(gate_any)]
        rows_for_size.append(row)
        for c, v in enumerate(row, 1):
            cell = ws.cell(rr, c, v)
            cell.font = BASE_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=isinstance(v, str) and len(str(v)) > 40)
        if gate_any:
            for c in range(1, len(headers) + 1):
                ws.cell(rr, c).fill = GATE_FILL
        rr += 1
    _autosize(ws, headers, rows_for_size)
    # APEX row
    rr += 1
    ws.cell(rr, 1, "APEX gcp_score = SUM(weight_in_apex x aggregate_score)").font = KEY_FONT
    ac = ws.cell(rr, 6, apex_str)
    ac.font = KEY_FONT
    ac.fill = APEX_FILL
    # Table B: per-point block scores
    tb = rr + 2
    ws.cell(tb, 1, "Per-point block scores").font = KEY_FONT
    tb += 1
    pheaders = ["point_id", "device_role", "device_type"]
    for bid in block_order:
        pheaders += [f"{bid}|score", f"{bid}|gate"]
    _header_row(ws, tb, pheaders)
    pr = tb + 1
    for pid in gcp_pids:
        ppb = ppb_by_pid.get(pid, {})
        row = [pid, ppb.get("device_role"), ppb.get("device_type")]
        gates = []
        for bid in block_order:
            bsc = ppb.get("block_scores", {}).get(bid, {})
            row += [bsc.get("score"), _fmt(bsc.get("gate_triggered"))]
            gates.append(bool(bsc.get("gate_triggered")))
        for c, v in enumerate(row, 1):
            cell = ws.cell(pr, c, v)
            cell.font = BASE_FONT
            cell.border = BORDER
        # tint gate cells
        for gi, g in enumerate(gates):
            if g:
                ws.cell(pr, 3 + gi * 2 + 1).fill = GATE_FILL
                ws.cell(pr, 3 + gi * 2 + 2).fill = GATE_FILL
        pr += 1
    ws.freeze_panes = "A5"

    # ========================= 6_GCP_Score =========================
    ws = wb.create_sheet("6_GCP_Score")
    ws["A1"] = "Level 0 - GCP Score (apex)"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Source: {out['stage3_gcp_score']}"
    ws["A2"].font = SUB_FONT
    metrics = [
        ("gcp_score", apex_str),
        ("apex is null?", _fmt(is_null)),
        ("gcp_role_point_count", gcp_count),
        ("excluded CHECK_POINT-role points", ", ".join(cp_pids) or "(none)"),
        ("weighted score before global gate", _na(d["weighted_score_before_global_gate"], is_null)),
        ("global gate triggered", _fmt(gg["triggered"])),
        ("global gate condition", gg["condition_spec"]),
        ("global gate action", gg["action_spec"]),
        ("null_handling condition", nh["condition_spec"]),
        ("no_gcp_role_points", _fmt(nh["no_gcp_role_points"])),
        ("apex formula", d["apex_formula_spec"]),
        ("apex weight-sum audit ok", _fmt(weight_ok)),
    ]
    rr = 4
    for name, val in metrics:
        ws.cell(rr, 1, name).font = KEY_FONT
        cc = ws.cell(rr, 2, val)
        cc.font = BASE_FONT
        cc.alignment = Alignment(wrap_text=True, vertical="top")
        for c in (1, 2):
            ws.cell(rr, c).border = BORDER
        if name == "gcp_score":
            cc.fill = GATE_FILL if is_null else APEX_FILL
        if name == "global gate triggered" and gg["triggered"]:
            cc.fill = GATE_FILL
        rr += 1
    # coverage_gate_fired_by_point table
    rr += 1
    ws.cell(rr, 1, "coverage_gate_fired_by_point").font = KEY_FONT
    rr += 1
    _header_row(ws, rr, ["point_id", "coverage gate fired"])
    rr += 1
    for pid, fired in (gg.get("coverage_gate_fired_by_point") or {}).items():
        ws.cell(rr, 1, pid).font = BASE_FONT
        cc = ws.cell(rr, 2, "GATE FIRED" if fired else "ok")
        cc.font = BASE_FONT
        if fired:
            cc.fill = GATE_FILL
        for c in (1, 2):
            ws.cell(rr, c).border = BORDER
        rr += 1
    # contributions LIST
    rr += 1
    ws.cell(rr, 1, "Block contributions (spec-formula order)").font = KEY_FONT
    rr += 1
    cheaders = ["block_id", "block_name", "weight_in_apex", "block_aggregate_score", "contribution"]
    _header_row(ws, rr, cheaders)
    rr += 1
    for con in d["contributions"]:
        for c, v in enumerate([con["block_id"], con["block_name"], con["weight_in_apex"],
                               _na(con["block_aggregate_score"], is_null),
                               _na(con["contribution"], is_null)], 1):
            cell = ws.cell(rr, c, v)
            cell.font = BASE_FONT
            cell.border = BORDER
        rr += 1
    # flags table
    rr += 1
    ws.cell(rr, 1, f"Flags raised - {sd['total_flags_aggregated']} aggregated "
                   f"(severity {_fmt(d['flags_by_severity']) or 'none'})").font = KEY_FONT
    rr += 1
    fheaders = ["flag_id", "flag_name", "severity", "origin_stage", "origin_point",
                "origin_block", "origin_indicator", "condition_value"]
    _header_row(ws, rr, fheaders)
    rr += 1
    flag_rows = d["all_flags_aggregated"] or [{"flag_id": "(none)",
               "flag_name": "no flags raised", "severity": "", "_origin_stage": "",
               "_origin_point": "", "_origin_block": "", "_origin_indicator": "", "condition_value": ""}]
    for f in flag_rows:
        vals = [f.get("flag_id"), f.get("flag_name"), f.get("severity"), f.get("_origin_stage"),
                f.get("_origin_point"), f.get("_origin_block"), f.get("_origin_indicator"),
                _fmt(f.get("condition_value"))]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(rr, c, v)
            cell.font = BASE_FONT
            cell.fill = FLAG_FILL if d["all_flags_aggregated"] else NA_FILL
            cell.border = BORDER
        rr += 1
    # handoff table / placeholder
    rr += 1
    ws.cell(rr, 1, "_handoff_crossdoc_candidates").font = KEY_FONT
    rr += 1
    if has_handoff and d.get("_handoff_crossdoc_candidates"):
        for h in d["_handoff_crossdoc_candidates"]:
            cc = ws.cell(rr, 1, _fmt(h, 400))
            cc.font = BASE_FONT
            cc.fill = HANDOFF_FILL
            rr += 1
    else:
        ws.cell(rr, 1, "(not currently emitted by GCP)").font = SUB_FONT
        rr += 1
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 42
    for c in range(3, len(fheaders) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 20

    # ========================= 8_Recommendations =========================
    ws = wb.create_sheet("8_Recommendations")
    if rec:
        write_recommendations(ws, rec)            # populated path (wired in Deliverable 4)
    else:
        write_recommendations_placeholder(ws)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    got = {a: len(spec[a]) for a in ["source_files", "source_fields", "derived_fields",
                                     "indicators", "building_blocks"]}
    headline = {"gcp_score": apex_str, "gcp_role_point_count": gcp_count,
                "gcp_points": len(gcp_pids), "cp_excluded": len(cp_pids),
                "global_gate_triggered": gg["triggered"], "is_null": is_null}
    return out_path, got, headline


def _na(v, is_null):
    return "N/A" if is_null else v


def _summary_rec_callout(ws, r, rec):
    ws.cell(r, 1, "RECOMMENDATION").font = Font(name=FONT, bold=True, size=12, color="1F4E78")
    r += 1
    decision = rec.get("decision") or "unable_to_assess"
    ws.cell(r, 1, "decision").font = KEY_FONT
    dc = ws.cell(r, 2, decision.upper().replace("_", " "))
    dc.font = Font(name=FONT, bold=True, size=13, color=DECISION_FG.get(decision, "000000"))
    dc.fill = DECISION_FILLS.get(decision, NA_FILL)
    r += 1
    ss = rec.get("subsystem_summary", {})
    pts = (f"good {ss.get('good_points', 0)} / review {ss.get('review_points', 0)} / "
           f"resurvey {ss.get('resurvey_points', 0)}  (n={ss.get('n_points', 0)})")
    rows = [("tier", rec.get("tier_interpretation")), ("rationale", rec.get("decision_rationale")),
            ("points by decision", pts), ("library version", rec.get("library_version"))]
    if rec.get("_spec_version_note"):
        rows.append(("spec/library note", rec["_spec_version_note"]))
    for k, v in rows:
        ws.cell(r, 1, k).font = KEY_FONT
        ws.cell(r, 2, v).font = BASE_FONT
        ws.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    hg = ss.get("hard_gates_fired_by_point") or {}
    ws.cell(r, 1, "hard_gates_fired_by_point").font = KEY_FONT
    cc = ws.cell(r, 2, _fmt(hg) if hg else "(none)")
    cc.font = BASE_FONT
    if hg:
        cc.fill = GATE_FILL
    r += 1
    return r


def write_recommendations_placeholder(ws):
    ws["A1"] = "8_Recommendations - PLACEHOLDER"
    ws["A1"].font = Font(name=FONT, bold=True, size=14, color="9C0006")
    ws["A1"].fill = HANDOFF_FILL
    ws["A3"] = "Recommendations have not been computed yet (outputs/07_recommendations.json not found)."
    ws["A3"].font = BASE_FONT
    ws["A5"] = "To populate this sheet (Deliverable 4):"
    ws["A5"].font = KEY_FONT
    ws["A6"] = "  /opt/anaconda3/bin/python3 scripts/compute_recommendations.py paths.json"
    ws["A6"].font = Font(name="Courier New", size=10)
    ws["A7"] = "  /opt/anaconda3/bin/python3 scripts/export_results_xlsx.py paths.json   (re-run to render)"
    ws["A7"].font = Font(name="Courier New", size=10)
    ws["A9"] = ("Once wired, this sheet shows: the chain decision callout, subsystem summary "
                "(good/review/resurvey point counts + hard_gates_fired_by_point), an indicator "
                "rollup, and a per-point fan-out of indicator recommendations.")
    ws["A9"].font = SUB_FONT
    ws["A9"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 110


def write_recommendations(ws, rec):
    """Full per-point recommendations sheet (Deliverable 4 STEP D)."""
    NCOL = 9
    decision = rec.get("decision") or "unable_to_assess"
    tc = ws.cell(1, 1, "8_Recommendations - GCP per-point (Tier 2 library)")
    tc.font = Font(name=FONT, bold=True, size=13, color="FFFFFF")
    tc.fill = HDR_FILL
    tc.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOL)
    ws.row_dimensions[1].height = 22

    rr = 3
    ws.cell(rr, 1, "Decision").font = KEY_FONT
    dc = ws.cell(rr, 2, decision.upper().replace("_", " "))
    dc.font = Font(name=FONT, bold=True, size=16, color=DECISION_FG.get(decision, "000000"))
    dc.fill = DECISION_FILLS.get(decision, NA_FILL)
    dc.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=4)
    ws.row_dimensions[rr].height = 30
    rr += 1

    def kv(label, val):
        nonlocal rr
        ws.cell(rr, 1, label).font = KEY_FONT
        c = ws.cell(rr, 2, _fmt(val) if isinstance(val, (dict, list)) else val)
        c.font = BASE_FONT
        c.alignment = Alignment(wrap_text=True, vertical="top")
        rr += 1

    kv("Apex score", "null" if rec.get("apex_score") is None else rec.get("apex_score"))
    kv("Tier interpretation", rec.get("tier_interpretation"))
    kv("Rationale", rec.get("decision_rationale"))
    kv("Library version", rec.get("library_version"))
    kv("Pipeline spec", rec.get("spec_version"))
    if rec.get("_spec_version_note"):
        kv("Spec/library note", rec["_spec_version_note"])
    kv("no_gcp_role_points", _fmt((rec.get("null_handling") or {}).get("no_gcp_role_points")))
    rr += 1

    # Subsystem summary
    ws.cell(rr, 1, "Subsystem summary").font = Font(name=FONT, bold=True, size=11, color="1F4E78")
    rr += 1
    _header_row(ws, rr, ["n_points", "good_points", "review_points", "resurvey_points",
                         "hard_gates_fired_by_point"])
    rr += 1
    ss = rec.get("subsystem_summary", {})
    hg = ss.get("hard_gates_fired_by_point") or {}
    for c, v in enumerate([ss.get("n_points", 0), ss.get("good_points", 0), ss.get("review_points", 0),
                           ss.get("resurvey_points", 0), _fmt(hg) if hg else "—"], 1):
        cell = ws.cell(rr, c, v)
        cell.font = BASE_FONT
        cell.border = BORDER
        if hg:
            cell.fill = GATE_FILL
    rr += 2

    # Indicator rollup
    ws.cell(rr, 1, "Indicator rollup (worst across points)").font = Font(name=FONT, bold=True, size=11, color="1F4E78")
    rr += 1
    _header_row(ws, rr, ["indicator_id", "worst_score", "worst_level", "n_points_in_review",
                         "n_points_in_resurvey", "hard_gate_points"])
    rr += 1
    for iid in sorted(rec.get("indicator_rollup", {})):
        ro = rec["indicator_rollup"][iid]
        lvl = ro.get("worst_level_across_points", "good")
        vals = [iid, ro.get("worst_score_across_points"), lvl, ro.get("n_points_in_review", 0),
                ro.get("n_points_in_resurvey", 0), ", ".join(ro.get("hard_gate_points") or []) or "—"]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(rr, c, v)
            cell.font = BASE_FONT
            cell.border = BORDER
            cell.fill = LEVEL_FILLS.get(lvl, NA_FILL)
        rr += 1
    rr += 1

    # Per-point fan-out
    if not rec.get("points"):
        ws.cell(rr, 1, "No GCP-role points — unable to assess (per-point detail empty).").font = SUB_FONT
        rr += 1
    for p in rec.get("points", []):
        hdr = ws.cell(rr, 1, f"Point {p['point_id']} ({p.get('device_role')}) — "
                             f"point_decision: {(p.get('point_decision') or '').upper()}")
        hdr.font = Font(name=FONT, bold=True, size=11, color="1F4E78")
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=NCOL)
        rr += 1
        _header_row(ws, rr, ["indicator_id", "name", "block", "score", "level", "band_label",
                             "recommendation_text", "actions", "hard_gate_fired"])
        rr += 1
        for ind in p.get("indicators", []):
            band = ind.get("matched_band") or {}
            level = band.get("level") or "unknown"
            is_good = level == "good"
            text = ind.get("verified_statement") if is_good else (ind.get("impact") or "")
            actions = ind.get("actions") or []
            actions_str = "• " + "\n• ".join(actions) if actions else ""
            fill = LEVEL_FILLS.get(level, NA_FILL)
            vals = [ind.get("indicator_id"), ind.get("name"), ind.get("block"), ind.get("score"),
                    level, band.get("label"), text, actions_str, _fmt(ind.get("hard_gate_fired"))]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(rr, c, _fmt(v) if isinstance(v, (dict, list)) else v)
                cell.font = BASE_FONT
                cell.fill = fill
                cell.border = BORDER
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.row_dimensions[rr].height = 56 if ((text and not is_good) or actions_str) else 22
            rr += 1
        rr += 1

    # Caveats
    caveats = rec.get("_caveats") or []
    ws.cell(rr, 1, f"Caveats ({len(caveats)})").font = Font(name=FONT, bold=True, size=11, color="1F4E78")
    rr += 1
    if caveats:
        _header_row(ws, rr, ["point_id", "indicator_id", "code", "detail"])
        rr += 1
        for cav in caveats:
            vals = [cav.get("point_id"), cav.get("indicator_id"), cav.get("code"),
                    _fmt({k: v for k, v in cav.items() if k not in ("point_id", "indicator_id", "code")})]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(rr, c, v)
                cell.font = BASE_FONT
                cell.border = BORDER
                cell.fill = FLAG_FILL
            rr += 1
    else:
        ws.cell(rr, 1, "No caveats — every (point, indicator) score landed in a library band cleanly.").font = BASE_FONT
        rr += 1

    for i, w in enumerate([22, 24, 22, 9, 11, 38, 60, 56, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"


def main(argv=None):
    ap = argparse.ArgumentParser(description="Export GCP pipeline results to a single xlsx")
    ap.add_argument("config", nargs="?", default="paths.json")
    ap.add_argument("out", nargs="?", default="outputs/gcp_results.xlsx")
    args = ap.parse_args(argv)
    config_path = Path(args.config).resolve()
    config = json.loads(config_path.read_text())
    root = config_path.parent
    rec_path = root / "outputs" / "07_recommendations.json"
    rec = json.loads(rec_path.read_text()) if rec_path.exists() else None
    out, counts, headline = build(root, config, root / args.out, rec=rec)
    try:
        shown = out.relative_to(root)
    except ValueError:
        shown = out
    print(f"Wrote {shown}  (sheets: Summary + 1-6 + 8_Recommendations"
          f"{' [placeholder]' if rec is None else ''}, no 5b)")
    print(f"  spec counts: {counts}")
    print(f"  headline: gcp_score={headline['gcp_score']}  "
          f"gcp_role_point_count={headline['gcp_role_point_count']}  "
          f"(gcp_points={headline['gcp_points']}, cp_excluded={headline['cp_excluded']}, "
          f"global_gate_triggered={headline['global_gate_triggered']}, null={headline['is_null']})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
